#!/usr/bin/env python3
"""
Profile a customer automation/use-case inventory for UiPath agentic expansion planning.

Inputs: .csv, .tsv, .xlsx, or .xlsm inventory files.
Outputs: inventory_profile.json and inventory_profile.md in the selected output directory.

This script is intentionally descriptive rather than prescriptive: it identifies inventory
structure, candidate columns, status distribution, owner/department density, value/volume
fields, missingness, duplicate names, and row samples for analyst review. It does not make
final recommendations by itself.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - only needed for xlsx input
    load_workbook = None


COLUMN_KEYWORDS: Dict[str, List[str]] = {
    "use_case_name": [
        "use case",
        "usecase",
        "automation name",
        "process name",
        "opportunity name",
        "idea name",
        "project name",
        "bot name",
        "solution name",
        "title",
        "name",
    ],
    "description": [
        "description",
        "business problem",
        "problem statement",
        "process description",
        "scope",
        "summary",
        "details",
        "manual process",
    ],
    "status": [
        "status",
        "stage",
        "phase",
        "state",
        "lifecycle",
        "deployment status",
        "project status",
    ],
    "department": [
        "department",
        "business unit",
        "agency",
        "division",
        "function",
        "organization",
        "org",
        "area",
        "team",
        "process owner group",
    ],
    "owner": [
        "owner",
        "process owner",
        "business owner",
        "sme",
        "sponsor",
        "requestor",
        "requester",
        "contact",
    ],
    "systems": [
        "application",
        "applications",
        "system",
        "systems",
        "platform",
        "erp",
        "source system",
        "target system",
        "technology",
    ],
    "volume": [
        "volume",
        "annual volume",
        "yearly volume",
        "monthly volume",
        "weekly volume",
        "transactions",
        "transaction count",
        "cases",
        "requests",
        "items",
    ],
    "weekly_volume": [
        "weekly volume",
        "per week",
        "weekly transactions",
        "weekly cases",
        "weekly requests",
    ],
    "annual_volume": [
        "annual volume",
        "yearly volume",
        "annual transactions",
        "annual cases",
        "annual requests",
        "per year",
    ],
    "handling_time": [
        "handling time",
        "average handling time",
        "average handling minutes",
        "avg handling minutes",
        "average handle time",
        "aht",
        "minutes per transaction",
        "hours per transaction",
        "manual time",
        "average processing minutes",
        "processing time",
        "time per case",
    ],
    "hours_saved": [
        "hours saved",
        "annual hours",
        "hours avoided",
        "capacity saved",
        "fte saved",
        "fte",
        "savings hours",
    ],
    "value": [
        "value",
        "savings",
        "cost savings",
        "benefit",
        "annual benefit",
        "roi",
        "return",
        "dollars",
        "usd",
    ],
    "priority": [
        "priority",
        "rank",
        "score",
        "business value",
        "impact",
        "complexity",
        "feasibility",
    ],
    "date": [
        "created",
        "submitted",
        "date",
        "go live",
        "golive",
        "deployment date",
        "last updated",
        "modified",
    ],
}

PRODUCTION_TERMS = [
    "production",
    "prod",
    "live",
    "deployed",
    "implemented",
    "in use",
    "active",
    "complete",
    "completed",
    "operational",
]
PIPELINE_TERMS = [
    "pipeline",
    "in progress",
    "development",
    "dev",
    "build",
    "testing",
    "uat",
    "pilot",
    "poc",
    "approved",
    "planned",
    "backlog",
    "assessment",
]
IDEA_TERMS = ["idea", "candidate", "intake", "submitted", "requested", "opportunity", "concept"]
EXCLUDED_TERMS = [
    "paused",
    "retired",
    "decommissioned",
    "sunset",
    "cancelled",
    "canceled",
    "abandoned",
    "rejected",
    "declined",
    "not approved",
    "duplicate",
    "duplicated",
    "archived",
    "on hold",
    "hold",
    "suspended",
    "deferred",
    "not started",
    "withdrawn",
]

PAUSED_TERMS = ["paused", "on hold", "hold", "suspended", "deferred"]
RETIRED_TERMS = ["retired", "decommissioned", "archived", "sunset"]
CANCELLED_TERMS = ["cancelled", "canceled", "withdrawn", "abandoned"]
REJECTED_TERMS = ["rejected", "declined", "not approved"]
DUPLICATE_TERMS = ["duplicate", "duplicated"]
NON_PRODUCTION_TERMS = [
    "not deployed",
    "never deployed",
    "never live",
    "not live",
    "not in production",
    "pre production",
    "preproduction",
    "not implemented",
]

STOP_WORDS = {
    "the",
    "and",
    "for",
    "with",
    "from",
    "into",
    "onto",
    "that",
    "this",
    "use",
    "case",
    "automation",
    "automated",
    "process",
    "bot",
    "robot",
    "rpa",
    "data",
    "report",
    "reports",
    "request",
    "requests",
    "review",
    "update",
    "create",
    "system",
    "manual",
    "task",
    "workflow",
}

PROFILE_SCHEMA_VERSION = "1.1"


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_text(value: Any) -> str:
    text = clean_cell(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_inventory_id(row: Dict[str, Any]) -> str:
    """Create a stable ID from source sheet and physical row number."""

    sheet = clean_cell(row.get("__sheet")) or "sheet"
    sheet_slug = re.sub(r"[^A-Z0-9]+", "-", sheet.upper()).strip("-") or "SHEET"
    row_number = row.get("__row_number")
    if isinstance(row_number, bool) or not isinstance(row_number, int) or row_number < 1:
        raise ValueError("inventory row is missing a positive __row_number")
    return f"INV-{sheet_slug[:48]}-R{row_number:05d}"


def nonempty(value: Any) -> bool:
    return clean_cell(value) != ""


def score_header(header: str, keywords: Iterable[str]) -> int:
    h = normalize_text(header)
    if not h:
        return 0
    best = 0
    for keyword in keywords:
        k = normalize_text(keyword)
        if not k:
            continue
        if h == k:
            best = max(best, 10)
        elif h.endswith(" " + k) or h.startswith(k + " "):
            best = max(best, 8)
        elif k in h:
            best = max(best, 6)
        elif all(part in h.split() for part in k.split()):
            best = max(best, 4)
    return best


def dedupe_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = defaultdict(int)
    result: List[str] = []
    for index, header in enumerate(headers, start=1):
        base = clean_cell(header) or f"column_{index}"
        seen[base] += 1
        if seen[base] == 1:
            result.append(base)
        else:
            result.append(f"{base}_{seen[base]}")
    return result


def sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        return dialect.delimiter
    except Exception:
        return "\t" if "\t" in sample else ","


def read_csv_inventory(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    last_error: Optional[Exception] = None
    for encoding in encodings:
        try:
            sample = path.read_text(encoding=encoding, errors="replace")[:4096]
            delimiter = sniff_delimiter(sample)
            rows: List[Dict[str, Any]] = []
            with path.open("r", encoding=encoding, errors="replace", newline="") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                raw_rows = list(reader)
            if not raw_rows:
                return {path.stem: []}
            header_index = find_header_row(raw_rows[:25])
            headers = dedupe_headers([clean_cell(x) for x in raw_rows[header_index]])
            for row_number, raw in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
                if not any(nonempty(x) for x in raw):
                    continue
                padded = raw + [""] * max(0, len(headers) - len(raw))
                row = {headers[i]: clean_cell(padded[i]) if i < len(padded) else "" for i in range(len(headers))}
                row["__row_number"] = row_number
                row["__sheet"] = path.stem
                rows.append(row)
            return {path.stem: rows}
        except Exception as exc:  # pragma: no cover - fallback path
            last_error = exc
    raise RuntimeError(f"could not read csv file: {last_error}")


def find_header_row(raw_rows: List[Iterable[Any]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(raw_rows):
        cells = [clean_cell(x) for x in row]
        filled = [x for x in cells if x]
        if len(filled) < 2:
            continue
        keyword_hits = 0
        for cell in filled:
            cell_norm = normalize_text(cell)
            for keywords in COLUMN_KEYWORDS.values():
                if score_header(cell_norm, keywords) >= 6:
                    keyword_hits += 1
                    break
        score = len(filled) + (keyword_hits * 4)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def read_xlsx_inventory(path: Path, selected_sheet: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for .xlsx/.xlsm input but is not installed")
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet_names = [selected_sheet] if selected_sheet else workbook.sheetnames
    sheets: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"sheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        raw_rows = [
            [clean_cell(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        if not any(any(values) for values in raw_rows):
            sheets[sheet_name] = []
            continue
        header_index = find_header_row(raw_rows[:25])
        headers = dedupe_headers(raw_rows[header_index])
        rows: List[Dict[str, Any]] = []
        for absolute_index, raw in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
            if not any(nonempty(x) for x in raw):
                continue
            padded = raw + [""] * max(0, len(headers) - len(raw))
            row = {headers[i]: clean_cell(padded[i]) if i < len(padded) else "" for i in range(len(headers))}
            row["__row_number"] = absolute_index
            row["__sheet"] = sheet_name
            rows.append(row)
        sheets[sheet_name] = rows
    return sheets


def load_inventory(path: Path, sheet: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return read_csv_inventory(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_inventory(path, selected_sheet=sheet)
    raise RuntimeError(f"unsupported file type: {suffix}. use .csv, .tsv, .xlsx, or .xlsm")


def get_headers(rows: List[Dict[str, Any]]) -> List[str]:
    ordered: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key.startswith("__"):
                continue
            if key not in seen:
                seen.add(key)
                ordered.append(key)
    return ordered


def guess_columns(headers: List[str]) -> Dict[str, Dict[str, Any]]:
    guesses: Dict[str, Dict[str, Any]] = {}
    for field_type, keywords in COLUMN_KEYWORDS.items():
        scored = []
        for header in headers:
            score = score_header(header, keywords)
            if score > 0:
                scored.append({"column": header, "score": score})
        scored.sort(key=lambda item: (-item["score"], item["column"].lower()))
        if scored:
            guesses[field_type] = {
                "best": scored[0]["column"],
                "score": scored[0]["score"],
                "alternates": [item["column"] for item in scored[1:5]],
            }
    return guesses


def parse_number(value: Any) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
    text = text.replace(",", "")
    text = text.replace("$", "")
    text = text.replace("%", "")
    match = re.search(
        r"-?\d+(?:\.\d+)?\s*(k|m|b|thousand|million|billion)?\b",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None
    try:
        numeric_match = re.search(r"-?\d+(?:\.\d+)?", match.group(0))
        if numeric_match is None:
            return None
        number = float(numeric_match.group(0))
        suffix = (match.group(1) or "").casefold()
        multiplier = {
            "": 1,
            "k": 1_000,
            "thousand": 1_000,
            "m": 1_000_000,
            "million": 1_000_000,
            "b": 1_000_000_000,
            "billion": 1_000_000_000,
        }[suffix]
        number *= multiplier
        return -number if negative else number
    except (KeyError, ValueError):
        return None


def parse_inventory_date(value: Any) -> Optional[str]:
    """Normalize an unambiguous inventory date to ISO YYYY-MM-DD."""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_cell(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        pass
    for pattern in ("%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def summarize_source_dates(
    rows: List[Dict[str, Any]],
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
) -> Dict[str, Any]:
    columns: List[str] = []
    raw_values: List[str] = []
    parsed_values: List[str] = []
    for sheet_name, mapping in sheet_field_mappings.items():
        column = mapping.get("date")
        if column and column not in columns:
            columns.append(column)
        for row in rows:
            if row.get("__sheet") != sheet_name or not column:
                continue
            raw = clean_cell(row.get(column))
            if not raw:
                continue
            raw_values.append(raw)
            parsed = parse_inventory_date(row.get(column))
            if parsed:
                parsed_values.append(parsed)
    return {
        "columns": columns,
        "nonblank_count": len(raw_values),
        "valid_count": len(parsed_values),
        "invalid_count": len(raw_values) - len(parsed_values),
        "earliest_date": min(parsed_values) if parsed_values else None,
        "latest_date": max(parsed_values) if parsed_values else None,
    }


def numeric_profile(rows: List[Dict[str, Any]], headers: List[str]) -> Dict[str, Dict[str, Any]]:
    profiles: Dict[str, Dict[str, Any]] = {}
    for header in headers:
        values = [parse_number(row.get(header)) for row in rows if nonempty(row.get(header))]
        values = [value for value in values if value is not None]
        nonblank = sum(1 for row in rows if nonempty(row.get(header)))
        if not values or nonblank == 0:
            continue
        coverage = len(values) / max(nonblank, 1)
        if len(values) >= 3 and coverage >= 0.5:
            sorted_values = sorted(values)
            profiles[header] = {
                "count": len(values),
                "min": sorted_values[0],
                "median": sorted_values[len(sorted_values) // 2],
                "max": sorted_values[-1],
                "sum": round(sum(values), 2),
                "parseable_coverage_of_nonblank": round(coverage, 3),
            }
    return profiles


def has_status_term(text: str, terms: Iterable[str]) -> bool:
    padded = f" {text} "
    return any(f" {normalize_text(term)} " in padded for term in terms)


def classify_status(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return "unknown"
    if has_status_term(text, EXCLUDED_TERMS):
        return "excluded"
    if has_status_term(text, NON_PRODUCTION_TERMS):
        return "idea" if has_status_term(text, IDEA_TERMS) else "pipeline"
    if has_status_term(text, PRODUCTION_TERMS):
        return "production"
    if has_status_term(text, PIPELINE_TERMS):
        return "pipeline"
    if has_status_term(text, IDEA_TERMS):
        return "idea"
    return "other"


def classify_lifecycle_status(value: Any) -> str:
    """Preserve customer-meaningful lifecycle states without weakening exclusion rules."""

    text = normalize_text(value)
    if not text or text in {"unknown", "n a", "na", "none", "not supplied"}:
        return "unknown"
    if has_status_term(text, DUPLICATE_TERMS):
        return "duplicate"
    if has_status_term(text, RETIRED_TERMS):
        return "retired"
    if has_status_term(text, CANCELLED_TERMS):
        return "cancelled"
    if has_status_term(text, REJECTED_TERMS):
        return "rejected"
    if has_status_term(text, PAUSED_TERMS):
        return "paused"
    if has_status_term(text, NON_PRODUCTION_TERMS):
        return "idea" if has_status_term(text, IDEA_TERMS) else "pipeline"
    if has_status_term(text, PRODUCTION_TERMS):
        return "deployed"
    if has_status_term(text, PIPELINE_TERMS) or has_status_term(text, ["not started"]):
        return "pipeline"
    if has_status_term(text, IDEA_TERMS):
        return "idea"
    return "other"


def coverage(rows: List[Dict[str, Any]], column: Optional[str]) -> Dict[str, Any]:
    if not column:
        return {"column": None, "nonblank": 0, "coverage_pct": 0.0}
    nonblank = sum(1 for row in rows if nonempty(row.get(column)))
    return {
        "column": column,
        "nonblank": nonblank,
        "coverage_pct": round((nonblank / len(rows) * 100.0) if rows else 0.0, 1),
    }


def top_counts(rows: List[Dict[str, Any]], column: Optional[str], limit: int = 20) -> List[Dict[str, Any]]:
    if not column:
        return []
    counter = Counter(clean_cell(row.get(column)) or "blank" for row in rows)
    return [{"value": key, "count": value} for key, value in counter.most_common(limit)]


def top_system_counts(
    rows: List[Dict[str, Any]], column: Optional[str], limit: int = 20
) -> List[Dict[str, Any]]:
    """Count individual systems when an inventory cell lists more than one."""

    if not column:
        return []
    counter: Counter[str] = Counter()
    labels: Dict[str, str] = {}
    for row in rows:
        raw_value = clean_cell(row.get(column))
        for value in re.split(r"\s*(?:[;,|\n]+|\s+/\s+)\s*", raw_value):
            label = value.strip()
            key = normalize_text(label)
            if not key:
                continue
            labels.setdefault(key, label)
            counter[key] += 1
    return [
        {"value": labels[key], "count": count}
        for key, count in counter.most_common(limit)
    ]


def mapped_column(
    row: Dict[str, Any],
    field: str,
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
) -> Optional[str]:
    sheet_name = str(row.get("__sheet", ""))
    return sheet_field_mappings.get(sheet_name, {}).get(field)


def mapped_value(
    row: Dict[str, Any],
    field: str,
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
) -> Any:
    column = mapped_column(row, field, sheet_field_mappings)
    return row.get(column) if column else ""


def mapped_metrics(
    row: Dict[str, Any],
    fields: List[str],
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
) -> List[Dict[str, Any]]:
    metrics: List[Dict[str, Any]] = []
    used_columns: set[str] = set()
    for field in fields:
        column = mapped_column(row, field, sheet_field_mappings)
        if not column or column in used_columns:
            continue
        used_columns.add(column)
        parsed = parse_number(row.get(column))
        if parsed is not None:
            metrics.append({"name": field, "value": parsed})
    return metrics


def mapped_coverage(
    rows: List[Dict[str, Any]],
    field: str,
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
    display_column: Optional[str],
) -> Dict[str, Any]:
    nonblank = sum(
        1 for row in rows if nonempty(mapped_value(row, field, sheet_field_mappings))
    )
    return {
        "column": display_column,
        "nonblank": nonblank,
        "coverage_pct": round((nonblank / len(rows) * 100.0) if rows else 0.0, 1),
    }


def mapped_top_counts(
    rows: List[Dict[str, Any]],
    field: str,
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
    limit: int = 20,
) -> List[Dict[str, Any]]:
    if not any(mapping.get(field) for mapping in sheet_field_mappings.values()):
        return []
    counter = Counter(
        clean_cell(mapped_value(row, field, sheet_field_mappings)) or "blank"
        for row in rows
    )
    return [{"value": key, "count": value} for key, value in counter.most_common(limit)]


def mapped_top_system_counts(
    rows: List[Dict[str, Any]],
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
    limit: int = 20,
) -> List[Dict[str, Any]]:
    counter: Counter[str] = Counter()
    labels: Dict[str, str] = {}
    for row in rows:
        raw_value = clean_cell(mapped_value(row, "systems", sheet_field_mappings))
        for value in re.split(r"\s*(?:[;,|\n]+|\s+/\s+)\s*", raw_value):
            label = value.strip()
            key = normalize_text(label)
            if not key:
                continue
            labels.setdefault(key, label)
            counter[key] += 1
    return [
        {"value": labels[key], "count": count}
        for key, count in counter.most_common(limit)
    ]


def safe_source_name(path: Path) -> str:
    """Return a display-safe basename without leaking the source directory."""

    name = re.sub(r"[\x00-\x1f\x7f]+", " ", path.name)
    name = re.sub(r"[<>:\"/\\|?*`]+", "-", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:255] or "inventory"


def extract_keywords(rows: List[Dict[str, Any]], columns: List[str], limit: int = 40) -> List[Dict[str, Any]]:
    counter: Counter[str] = Counter()
    for row in rows:
        text = " ".join(clean_cell(row.get(col)) for col in columns if col)
        for token in re.findall(r"[a-zA-Z][a-zA-Z0-9]{2,}", text.lower()):
            if token not in STOP_WORDS and len(token) > 2:
                counter[token] += 1
    return [{"term": term, "count": count} for term, count in counter.most_common(limit)]


def mapped_extract_keywords(
    rows: List[Dict[str, Any]],
    fields: List[str],
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
    limit: int = 40,
) -> List[Dict[str, Any]]:
    counter: Counter[str] = Counter()
    for row in rows:
        text = " ".join(
            clean_cell(mapped_value(row, field, sheet_field_mappings))
            for field in fields
        )
        for token in re.findall(r"[a-zA-Z][a-zA-Z0-9]{2,}", text.lower()):
            if token not in STOP_WORDS and len(token) > 2:
                counter[token] += 1
    return [{"term": term, "count": count} for term, count in counter.most_common(limit)]


def duplicate_name_groups(rows: List[Dict[str, Any]], name_column: Optional[str], limit: int = 30) -> List[Dict[str, Any]]:
    if not name_column:
        return []
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        raw_name = clean_cell(row.get(name_column))
        key = normalize_text(raw_name)
        if key:
            groups[key].append(row)
    duplicates = []
    for key, group in groups.items():
        if len(group) > 1:
            duplicates.append(
                {
                    "normalized_name": key,
                    "count": len(group),
                    "examples": [
                        {
                            "name": clean_cell(item.get(name_column)),
                            "sheet": item.get("__sheet"),
                            "row_number": item.get("__row_number"),
                        }
                        for item in group[:5]
                    ],
                }
            )
    duplicates.sort(key=lambda item: (-item["count"], item["normalized_name"]))
    return duplicates[:limit]


def mapped_duplicate_name_groups(
    rows: List[Dict[str, Any]],
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
    limit: int = 30,
) -> List[Dict[str, Any]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        raw_name = clean_cell(mapped_value(row, "use_case_name", sheet_field_mappings))
        key = normalize_text(raw_name)
        if key:
            groups[key].append(row)
    duplicates = []
    for key, group in groups.items():
        if len(group) > 1:
            duplicates.append(
                {
                    "normalized_name": key,
                    "count": len(group),
                    "examples": [
                        {
                            "name": clean_cell(
                                mapped_value(item, "use_case_name", sheet_field_mappings)
                            ),
                            "sheet": item.get("__sheet"),
                            "row_number": item.get("__row_number"),
                        }
                        for item in group[:5]
                    ],
                }
            )
    duplicates.sort(key=lambda item: (-item["count"], item["normalized_name"]))
    return duplicates[:limit]


def compact_row(row: Dict[str, Any], columns: List[str]) -> Dict[str, Any]:
    result = {
        "inventory_id": stable_inventory_id(row),
        "sheet": row.get("__sheet"),
        "row_number": row.get("__row_number"),
    }
    for column in columns:
        if column and column in row:
            value = clean_cell(row.get(column))
            if len(value) > 220:
                value = value[:217] + "..."
            result[column] = value
    return result


def compact_mapped_row(
    row: Dict[str, Any],
    fields: List[str],
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]],
) -> Dict[str, Any]:
    result = {
        "inventory_id": stable_inventory_id(row),
        "sheet": row.get("__sheet"),
        "row_number": row.get("__row_number"),
    }
    for field in fields:
        column = mapped_column(row, field, sheet_field_mappings)
        if not column:
            continue
        value = clean_cell(row.get(column))
        if len(value) > 220:
            value = value[:217] + "..."
        result[column] = value
    return result


def top_rows_by_metric(
    rows: List[Dict[str, Any]],
    metric_columns: List[str],
    context_columns: List[str],
    limit: int = 10,
    *,
    context_fields: Optional[List[str]] = None,
    sheet_field_mappings: Optional[Dict[str, Dict[str, Optional[str]]]] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    output: Dict[str, List[Dict[str, Any]]] = {}
    for column in metric_columns:
        scored = []
        for row in rows:
            number = parse_number(row.get(column))
            if number is None:
                continue
            scored.append((number, row))
        scored.sort(key=lambda pair: pair[0], reverse=True)
        if scored:
            output[column] = [
                {
                    "metric_value": number,
                    **(
                        compact_mapped_row(row, context_fields, sheet_field_mappings)
                        if context_fields is not None and sheet_field_mappings is not None
                        else compact_row(row, context_columns)
                    ),
                }
                for number, row in scored[:limit]
            ]
    return output


def build_profile(path: Path, sheets: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Any]:
    all_rows: List[Dict[str, Any]] = []
    sheet_profiles = []
    sheet_detected_columns: Dict[str, Dict[str, Dict[str, Any]]] = {}
    sheet_field_mappings: Dict[str, Dict[str, Optional[str]]] = {}
    for sheet_name, rows in sheets.items():
        headers = get_headers(rows)
        sheet_guesses = guess_columns(headers)
        sheet_detected_columns[sheet_name] = sheet_guesses
        sheet_field_mappings[sheet_name] = {
            field: sheet_guesses.get(field, {}).get("best")
            for field in COLUMN_KEYWORDS
        }
        all_rows.extend(rows)
        sheet_profiles.append(
            {
                "sheet": sheet_name,
                "rows": len(rows),
                "columns": headers,
                "column_count": len(headers),
            }
        )

    headers = get_headers(all_rows)
    generated_ids = [stable_inventory_id(row) for row in all_rows]
    duplicate_ids = sorted(
        item_id for item_id, count in Counter(generated_ids).items() if count > 1
    )
    if duplicate_ids:
        raise ValueError(
            "inventory ID collision after sheet-name normalization: "
            + ", ".join(duplicate_ids)
            + ". Rename colliding sheets and regenerate the profile."
        )
    guesses = guess_columns(headers)
    core_fields: Dict[str, Optional[str]] = {}
    for field in COLUMN_KEYWORDS:
        mapped_columns = []
        for sheet_name in sheets:
            column = sheet_field_mappings[sheet_name].get(field)
            if column and column not in mapped_columns:
                mapped_columns.append(column)
        core_fields[field] = mapped_columns[0] if mapped_columns else None

    field_available = {
        field: any(mapping.get(field) for mapping in sheet_field_mappings.values())
        for field in COLUMN_KEYWORDS
    }
    status_breakdown = (
        Counter(
            classify_status(mapped_value(row, "status", sheet_field_mappings))
            for row in all_rows
        )
        if field_available["status"]
        else Counter()
    )
    lifecycle_breakdown = (
        Counter(
            classify_lifecycle_status(mapped_value(row, "status", sheet_field_mappings))
            for row in all_rows
        )
        if field_available["status"]
        else Counter({"unknown": len(all_rows)})
    )
    raw_status_breakdown = mapped_top_counts(
        all_rows, "status", sheet_field_mappings, limit=30
    )
    source_date_summary = summarize_source_dates(all_rows, sheet_field_mappings)

    numeric = numeric_profile(all_rows, headers)
    detected_metric_columns = []
    for candidate in [
        core_fields["volume"],
        core_fields["weekly_volume"],
        core_fields["annual_volume"],
        core_fields["handling_time"],
        core_fields["hours_saved"],
        core_fields["value"],
        core_fields["priority"],
    ]:
        if candidate and candidate in numeric and candidate not in detected_metric_columns:
            detected_metric_columns.append(candidate)
    if len(detected_metric_columns) < 5:
        for column in numeric.keys():
            if column not in detected_metric_columns:
                detected_metric_columns.append(column)
            if len(detected_metric_columns) >= 8:
                break

    context_fields = [
        field
        for field in (
            "use_case_name",
            "description",
            "status",
            "department",
            "owner",
            "systems",
        )
        if field_available[field]
    ]
    context_columns = [core_fields[field] for field in context_fields if core_fields[field]]
    metric_fields: List[str] = []
    for field in (
        "annual_volume",
        "weekly_volume",
        "volume",
        "handling_time",
        "hours_saved",
        "value",
        "priority",
    ):
        if field_available[field]:
            metric_fields.append(field)

    required_for_full_quality = ["use_case_name", "description", "status", "department"]
    missing_core = [field for field in required_for_full_quality if not field_available[field]]
    weak_value_fields = not any(
        field_available[field]
        for field in [
            "volume",
            "weekly_volume",
            "annual_volume",
            "handling_time",
            "hours_saved",
            "value",
        ]
    )

    profile = {
        "schema_version": PROFILE_SCHEMA_VERSION,
        "metadata": {
            "source_file": safe_source_name(path),
            "source_name": safe_source_name(path),
            "source_sha256": sha256_file(path),
            "generated_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
            "sheet_count": len(sheets),
            "total_rows": len(all_rows),
            "total_columns": len(headers),
            "source_date_summary": source_date_summary,
        },
        "sheets": sheet_profiles,
        "detected_columns": guesses,
        "detected_columns_by_sheet": sheet_detected_columns,
        "core_field_mapping": core_fields,
        "sheet_field_mappings": sheet_field_mappings,
        "data_quality": {
            "missing_core_fields_for_full_quality": missing_core,
            "no_value_or_volume_fields_detected": weak_value_fields,
            "field_coverage": {
                field: mapped_coverage(
                    all_rows,
                    field,
                    sheet_field_mappings,
                    core_fields[field],
                )
                for field in core_fields
            },
            "duplicate_name_groups": mapped_duplicate_name_groups(
                all_rows, sheet_field_mappings
            ),
        },
        "status_summary": {
            "normalized_status_counts": dict(status_breakdown),
            "lifecycle_status_counts": dict(lifecycle_breakdown),
            "raw_status_top_values": raw_status_breakdown,
        },
        "owner_department_summary": {
            "department_top_values": mapped_top_counts(
                all_rows, "department", sheet_field_mappings
            ),
            "owner_top_values": mapped_top_counts(
                all_rows, "owner", sheet_field_mappings
            ),
        },
        "systems_summary": {
            "systems_top_values": mapped_top_system_counts(
                all_rows, sheet_field_mappings
            ),
        },
        "numeric_profiles": numeric,
        "text_patterns": {
            "frequent_terms_from_name_description": mapped_extract_keywords(
                all_rows,
                [field for field in ("use_case_name", "description") if field_available[field]],
                sheet_field_mappings,
            )
        },
        "top_rows_by_detected_metrics": top_rows_by_metric(
            all_rows,
            detected_metric_columns,
            context_columns,
            context_fields=context_fields,
            sheet_field_mappings=sheet_field_mappings,
        ),
        "representative_rows": [
            compact_mapped_row(row, context_fields, sheet_field_mappings)
            for row in all_rows[:15]
        ],
        "inventory_items": [
            {
                "inventory_id": stable_inventory_id(row),
                "name": clean_cell(mapped_value(row, "use_case_name", sheet_field_mappings)),
                "description": clean_cell(
                    mapped_value(row, "description", sheet_field_mappings)
                ),
                "status": (
                    classify_status(mapped_value(row, "status", sheet_field_mappings))
                    if field_available["status"]
                    else "unknown"
                ),
                "raw_status": clean_cell(mapped_value(row, "status", sheet_field_mappings)),
                "lifecycle_status": (
                    classify_lifecycle_status(
                        mapped_value(row, "status", sheet_field_mappings)
                    )
                    if field_available["status"]
                    else "unknown"
                ),
                "department": clean_cell(
                    mapped_value(row, "department", sheet_field_mappings)
                ),
                "owner": clean_cell(mapped_value(row, "owner", sheet_field_mappings)),
                "systems": clean_cell(mapped_value(row, "systems", sheet_field_mappings)),
                "source_date": parse_inventory_date(
                    mapped_value(row, "date", sheet_field_mappings)
                ),
                "sheet": row.get("__sheet"),
                "row_number": row.get("__row_number"),
                "metrics": mapped_metrics(row, metric_fields, sheet_field_mappings),
            }
            for row in all_rows
        ],
    }
    return profile


def markdown_table(rows: List[Dict[str, Any]], columns: List[str]) -> str:
    if not rows:
        return "No data detected.\n"
    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join(["---"] * len(columns)) + " |"
    lines = [header, separator]
    for row in rows:
        cells = []
        for column in columns:
            value = clean_cell(row.get(column, ""))
            value = value.replace("|", "\\|")
            cells.append(value)
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def profile_to_markdown(profile: Dict[str, Any]) -> str:
    meta = profile["metadata"]
    lines = [
        "# Inventory profile",
        "",
        f"Profile schema: `{profile['schema_version']}`",
        f"Source file: `{meta['source_file']}`",
        f"Source name: `{meta['source_name']}`",
        f"Source SHA-256: `{meta['source_sha256']}`",
        f"Generated UTC: `{meta['generated_at_utc']}`",
        f"Sheets: {meta['sheet_count']}",
        f"Rows: {meta['total_rows']}",
        f"Columns: {meta['total_columns']}",
        "Latest source record date: "
        + (meta.get("source_date_summary", {}).get("latest_date") or "not available"),
        "",
        "## Sheets",
    ]
    sheet_rows = [
        {"sheet": item["sheet"], "rows": str(item["rows"]), "columns": str(item["column_count"])}
        for item in profile["sheets"]
    ]
    lines.append(markdown_table(sheet_rows, ["sheet", "rows", "columns"]))

    lines.extend(["", "## Detected core field mapping"])
    mapping_rows = []
    for field, column in profile["core_field_mapping"].items():
        coverage_obj = profile["data_quality"]["field_coverage"].get(field, {})
        mapping_rows.append(
            {
                "field": field,
                "detected_column": column or "not detected",
                "coverage_pct": str(coverage_obj.get("coverage_pct", 0.0)),
            }
        )
    lines.append(markdown_table(mapping_rows, ["field", "detected_column", "coverage_pct"]))

    lines.extend(["", "## Data quality flags"])
    dq = profile["data_quality"]
    missing = dq["missing_core_fields_for_full_quality"]
    lines.append(f"- Missing core fields for full-quality output: {', '.join(missing) if missing else 'none detected'}")
    lines.append(f"- No value or volume fields detected: {dq['no_value_or_volume_fields_detected']}")
    lines.append(f"- Duplicate name groups detected: {len(dq['duplicate_name_groups'])}")

    lines.extend(["", "## Normalized status counts"])
    status_rows = [
        {"status_category": key, "count": str(value)}
        for key, value in sorted(profile["status_summary"]["normalized_status_counts"].items())
    ]
    lines.append(markdown_table(status_rows, ["status_category", "count"]))

    lines.extend(["", "## Detailed lifecycle status counts"])
    lifecycle_rows = [
        {"lifecycle_status": key, "count": str(value)}
        for key, value in sorted(profile["status_summary"]["lifecycle_status_counts"].items())
    ]
    lines.append(markdown_table(lifecycle_rows, ["lifecycle_status", "count"]))

    lines.extend(["", "## Top departments"])
    dept_rows = [
        {"department": item["value"], "count": str(item["count"])}
        for item in profile["owner_department_summary"]["department_top_values"][:15]
    ]
    lines.append(markdown_table(dept_rows, ["department", "count"]))

    lines.extend(["", "## Numeric fields"])
    numeric_rows = []
    for column, stats in profile["numeric_profiles"].items():
        numeric_rows.append(
            {
                "column": column,
                "count": str(stats["count"]),
                "median": str(stats["median"]),
                "max": str(stats["max"]),
                "sum": str(stats["sum"]),
            }
        )
    lines.append(markdown_table(numeric_rows, ["column", "count", "median", "max", "sum"]))

    lines.extend(["", "## Frequent terms from names and descriptions"])
    term_rows = [
        {"term": item["term"], "count": str(item["count"])}
        for item in profile["text_patterns"]["frequent_terms_from_name_description"][:25]
    ]
    lines.append(markdown_table(term_rows, ["term", "count"]))

    lines.extend(["", "## Representative rows"])
    rep_rows = profile["representative_rows"][:10]
    if rep_rows:
        columns = list(rep_rows[0].keys())
        lines.append(markdown_table(rep_rows, columns))
    else:
        lines.append("No representative rows available.\n")

    lines.extend(["", "## Inventory IDs"])
    id_rows = profile["inventory_items"][:25]
    if id_rows:
        lines.append(markdown_table(id_rows, ["inventory_id", "name", "status", "sheet", "row_number"]))
    else:
        lines.append("No inventory rows available.\n")

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Profile a UiPath customer use-case inventory.")
    parser.add_argument("--input", required=True, help="Path to .csv, .tsv, .xlsx, or .xlsm inventory file")
    parser.add_argument("--outdir", required=True, help="Directory where profile outputs should be written")
    parser.add_argument("--sheet", default=None, help="Optional worksheet name for Excel input")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    outdir = Path(args.outdir).expanduser().resolve()
    if not input_path.exists():
        print(f"input file not found: {input_path}", file=sys.stderr)
        return 2
    outdir.mkdir(parents=True, exist_ok=True)

    try:
        sheets = load_inventory(input_path, sheet=args.sheet)
        profile = build_profile(input_path, sheets)
    except Exception as exc:
        print(f"failed to profile inventory: {exc}", file=sys.stderr)
        return 1

    json_path = outdir / "inventory_profile.json"
    md_path = outdir / "inventory_profile.md"
    json_path.write_text(json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8")
    md_path.write_text(profile_to_markdown(profile), encoding="utf-8")
    print(f"wrote {json_path}")
    print(f"wrote {md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
