import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path


OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None
ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "research_row_sources.py"
FIXTURES = Path(__file__).resolve().parent / "fixtures"


def load_module():
    spec = importlib.util.spec_from_file_location("research_row_sources", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is not installed")
class ResearchRowSourcesTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.module = load_module()
        from openpyxl import Workbook, load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        cls.Workbook = Workbook
        cls.load_workbook = staticmethod(load_workbook)
        cls.DataValidation = DataValidation

    def write_shifted_main_workbook(
        self,
        path: Path,
        *,
        duplicate_account: bool = False,
        formula_in_ixp: bool = False,
    ) -> None:
        wb = self.Workbook()
        ws = wb.active
        ws.title = self.module.MAIN_SHEET
        shifted_headers = {
            2: "Cloud?",
            4: "Customer Name",
            6: "License Utilization",
            8: "AI Units Consumed",
            10: "Agent Units Purchased",
            12: "Test Suite Status",
            14: "DU/IXP Status",
            16: "Agents Status",
            18: "Adoption Segment",
            20: "FY27 Priorities",
            22: "Value Tracking",
            24: "Risk/Churn",
            26: "Evidence Notes",
        }
        for col, header in shifted_headers.items():
            ws.cell(12, col).value = header
        ws.cell(13, 4).value = "Department of Fixtures"
        ws.cell(13, 2).value = "-"
        ws.cell(13, 6).value = "Moderate"
        ws.cell(13, 8).value = "N"
        ws.cell(13, 14).value = '=IF(1=1,"PoC","")' if formula_in_ixp else None
        ws.cell(13, 27).value = "=1+1"
        if duplicate_account:
            ws.cell(14, 4).value = "Department-of-Fixtures"

        cloud_validation = self.DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
        ws.add_data_validation(cloud_validation)
        cloud_validation.add("B13:B100")

        current = wb.create_sheet("Current 2026-06-25")
        current.cell(3, 2).value = "Customer Name"
        current.cell(3, 4).value = "Notes"
        current.cell(3, 6).value = "Last Updated"
        current.cell(4, 2).value = "Department of Fixtures"
        current.cell(4, 4).value = "Document Understanding pilot for permit intake."
        current.cell(4, 6).value = "2026-06-25"

        stale = wb.create_sheet("Archive 2025-01-01")
        stale.append(["Account", "Notes", "Last Updated"])
        stale.append(["Department of Fixtures", "High risk and agentic exploration.", "2025-01-02"])
        wb.save(path)

    def write_source_workbook(self, path: Path, *, updated: str, notes: str) -> None:
        wb = self.Workbook()
        ws = wb.active
        ws.title = "Accounts"
        ws.cell(4, 2).value = "Customer Name"
        ws.cell(4, 4).value = "Current Platform"
        ws.cell(4, 6).value = "Notes"
        ws.cell(4, 8).value = "Last Modified"
        ws.cell(5, 2).value = "Department of Fixtures"
        ws.cell(5, 4).value = "Automation Cloud"
        ws.cell(5, 6).value = notes
        ws.cell(5, 8).value = updated
        wb.save(path)

    def copy_fixture(self, name: str, destination: Path, **replacements) -> dict:
        payload = json.loads((FIXTURES / name).read_text(encoding="utf-8"))
        payload.update(replacements)
        destination.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        return payload

    def write_empty_manifest(self, path: Path, *, version: str | None = None) -> None:
        payload = {
            "contract_version": version or self.module.SOURCE_MANIFEST_CONTRACT,
            "data_classification": "UiPath Confidential",
            "retention_until": "2026-12-31",
            "sources": [],
        }
        path.write_text(json.dumps(payload), encoding="utf-8")

    def run_cli(self, workbook: Path, manifest: Path | None, *extra: str) -> subprocess.CompletedProcess:
        command = [
            sys.executable,
            str(SCRIPT),
            "--workbook",
            str(workbook),
        ]
        if manifest is not None:
            command.extend(["--manifest", str(manifest)])
        command.extend(extra)
        return subprocess.run(command, capture_output=True, text=True, check=False)

    def test_shifted_schema_manifest_and_stale_evidence_are_routed_separately(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            current_source = tmp_path / "source-current.xlsx"
            stale_source = tmp_path / "source-stale.xlsx"
            manifest = tmp_path / "manifest.json"
            self.write_shifted_main_workbook(workbook)
            self.write_source_workbook(
                current_source,
                updated="2026-06-20",
                notes="Automation Cloud and using DU for permit intake.",
            )
            self.write_source_workbook(
                stale_source,
                updated="2025-01-01",
                notes="High risk, churn, and agentic exploration.",
            )
            self.copy_fixture("source-manifest-v1.json", manifest)

            result = self.run_cli(
                workbook,
                manifest,
                "--account",
                "Department of Fixtures",
                "--include-stale",
                "--as-of-date",
                "2026-07-10",
                "--format",
                "json",
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads(result.stdout)
            self.assertEqual(payload["contract_version"], self.module.OUTPUT_CONTRACT)
            self.assertEqual(payload["target_row"]["header_row"], 12)
            self.assertEqual(payload["target_row"]["account_column"], 4)
            self.assertEqual(payload["target_row"]["target_columns"]["Cloud Y/N"]["coordinate"], "B13")
            self.assertEqual(
                payload["target_row"]["target_columns"]["Bot/License Utilization"]["coordinate"],
                "F13",
            )
            fill_ids = {item["evidence_id"] for item in payload["evidence"]["fill_eligible_current"]}
            discovery_ids = {item["evidence_id"] for item in payload["evidence"]["discovery_leads"]}
            self.assertIn("current-source:Accounts:5", fill_ids)
            self.assertIn("stale-source:Accounts:5", discovery_ids)
            self.assertNotIn("stale-source:Accounts:5", fill_ids)
            self.assertTrue(
                all(item["freshness"]["status"] == "current" for item in payload["evidence"]["fill_eligible_current"])
            )
            self.assertTrue(
                all(
                    item["freshness"]["local_file_mtime_is_eligibility_evidence"] is False
                    for item in payload["evidence"]["fill_eligible_current"]
                )
            )
            self.assertIn("Cloud Y/N", payload["recommendation_leads"])
            self.assertIn("IXP Status", payload["recommendation_leads"])
            self.assertNotIn("At Risk/Churn Forecasted: Y/N", payload["recommendation_leads"])

    def test_stale_evidence_is_excluded_without_include_stale(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            current_source = tmp_path / "source-current.xlsx"
            stale_source = tmp_path / "source-stale.xlsx"
            manifest = tmp_path / "manifest.json"
            self.write_shifted_main_workbook(workbook)
            self.write_source_workbook(current_source, updated="2026-06-20", notes="Automation Cloud.")
            self.write_source_workbook(stale_source, updated="2025-01-01", notes="High risk.")
            self.copy_fixture("source-manifest-v1.json", manifest)

            result = self.run_cli(
                workbook,
                manifest,
                "--row",
                "13",
                "--as-of-date",
                "2026-07-10",
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads(result.stdout)
            discovery_ids = {item["evidence_id"] for item in payload["evidence"]["discovery_leads"]}
            excluded_ids = {item["evidence_id"] for item in payload["evidence"]["excluded"]}
            self.assertNotIn("stale-source:Accounts:5", discovery_ids)
            self.assertIn("stale-source:Accounts:5", excluded_ids)

    def test_ambiguous_exact_account_returns_candidates_and_nonzero_status(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            manifest = tmp_path / "manifest.json"
            self.write_shifted_main_workbook(workbook, duplicate_account=True)
            self.write_empty_manifest(manifest)

            result = self.run_cli(
                workbook,
                manifest,
                "--account",
                "Department of Fixtures",
                "--as-of-date",
                "2026-07-10",
            )

            self.assertEqual(result.returncode, 2)
            payload = json.loads(result.stderr)
            self.assertEqual(payload["error"]["code"], "ambiguous_account")
            self.assertEqual([item["row"] for item in payload["error"]["candidates"]], [13, 14])

    def test_near_account_name_is_returned_only_as_nonselecting_candidate(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            manifest = tmp_path / "manifest.json"
            self.write_shifted_main_workbook(workbook)
            wb = self.load_workbook(workbook)
            try:
                wb[self.module.MAIN_SHEET]["D13"] = "Office of Fixtures"
                wb.save(workbook)
            finally:
                wb.close()
            self.write_empty_manifest(manifest)

            result = self.run_cli(
                workbook,
                manifest,
                "--account",
                "Department of Fixtures",
                "--as-of-date",
                "2026-07-10",
            )

            self.assertEqual(result.returncode, 2)
            payload = json.loads(result.stderr)
            self.assertEqual(payload["error"]["code"], "account_not_found")
            self.assertEqual(payload["error"]["candidates"][0]["account"], "Office of Fixtures")

    def test_manifest_is_required_and_legacy_source_flags_fail_with_migration_guidance(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            self.write_shifted_main_workbook(workbook)

            missing_manifest = self.run_cli(
                workbook,
                None,
                "--row",
                "13",
                "--as-of-date",
                "2026-07-10",
            )
            self.assertEqual(missing_manifest.returncode, 2)
            self.assertEqual(json.loads(missing_manifest.stderr)["error"]["code"], "manifest_required")

            legacy = self.run_cli(
                workbook,
                None,
                "--row",
                "13",
                "--source",
                str(tmp_path / "legacy.xlsx"),
                "--sources-only",
            )
            self.assertEqual(legacy.returncode, 2)
            legacy_error = json.loads(legacy.stderr)["error"]
            self.assertEqual(legacy_error["code"], "legacy_source_flags_removed")
            self.assertIn("--manifest", legacy_error["message"])

    def test_unsupported_manifest_version_fails_closed(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            manifest = tmp_path / "manifest.json"
            self.write_shifted_main_workbook(workbook)
            self.write_empty_manifest(manifest, version="pubsec-big-rocks-row-research/source-manifest@2")

            result = self.run_cli(
                workbook,
                manifest,
                "--row",
                "13",
                "--as-of-date",
                "2026-07-10",
            )

            self.assertEqual(result.returncode, 2)
            self.assertEqual(json.loads(result.stderr)["error"]["code"], "unsupported_manifest_version")

    def test_preview_blocks_dropdown_formula_and_existing_value_updates(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            manifest = tmp_path / "manifest.json"
            updates = tmp_path / "invalid-updates.json"
            self.write_shifted_main_workbook(workbook, formula_in_ixp=True)
            self.write_empty_manifest(manifest)
            updates.write_text(
                json.dumps(
                    {
                        "contract_version": self.module.PROPOSED_UPDATES_CONTRACT,
                        "target": {"account": "Department of Fixtures", "row": 13},
                        "updates": [
                            {
                                "header": "Cloud Y/N",
                                "value": "Maybe",
                                "confidence": "High",
                                "evidence": ["manual:cloud"],
                            },
                            {
                                "header": "Bot/License Utilization",
                                "value": "High",
                                "confidence": "High",
                                "evidence": ["manual:utilization"],
                            },
                            {
                                "header": "IXP Status",
                                "value": "PoC",
                                "confidence": "Medium",
                                "evidence": ["manual:ixp"],
                            },
                        ],
                    }
                ),
                encoding="utf-8",
            )
            source_bytes = workbook.read_bytes()

            result = self.run_cli(
                workbook,
                manifest,
                "--row",
                "13",
                "--as-of-date",
                "2026-07-10",
                "--proposed-updates",
                str(updates),
            )

            self.assertEqual(result.returncode, 3, result.stderr)
            payload = json.loads(result.stdout)
            preview = payload["proposed_update_preview"]
            self.assertFalse(preview["valid"])
            by_header = {item["header"]: item for item in preview["proposed_changes"]}
            self.assertFalse(by_header["Cloud Y/N"]["checks"]["dropdown_valid"])
            self.assertFalse(
                by_header["Bot/License Utilization"]["checks"]["existing_value_blank_or_placeholder"]
            )
            self.assertFalse(by_header["IXP Status"]["checks"]["formula_safe"])
            self.assertFalse(by_header["IXP Status"]["checks"]["evidence_ids_fill_eligible"])
            self.assertEqual(workbook.read_bytes(), source_bytes)

    def test_write_copy_preserves_source_formulas_values_validations_and_sets_red_font(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            current_source = tmp_path / "source-current.xlsx"
            stale_source = tmp_path / "source-stale.xlsx"
            manifest = tmp_path / "manifest.json"
            updates = tmp_path / "updates.json"
            output = tmp_path / "big-rocks-updated.xlsx"
            self.write_shifted_main_workbook(workbook)
            self.write_source_workbook(current_source, updated="2026-06-20", notes="Automation Cloud.")
            self.write_source_workbook(stale_source, updated="2025-01-01", notes="Old evidence.")
            self.copy_fixture("source-manifest-v1.json", manifest)
            self.copy_fixture("proposed-updates-v1.json", updates)
            source_bytes = workbook.read_bytes()

            result = self.run_cli(
                workbook,
                manifest,
                "--row",
                "13",
                "--as-of-date",
                "2026-07-10",
                "--proposed-updates",
                str(updates),
                "--write-copy",
                str(output),
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            payload = json.loads(result.stdout)
            self.assertTrue(payload["proposed_update_preview"]["valid"])
            self.assertTrue(payload["write_copy"]["verified"])
            self.assertTrue(all(payload["write_copy"]["checks"].values()))
            self.assertEqual(workbook.read_bytes(), source_bytes)
            self.assertTrue(output.exists())

            source_wb = self.load_workbook(workbook, data_only=False)
            output_wb = self.load_workbook(output, data_only=False)
            try:
                source_ws = source_wb[self.module.MAIN_SHEET]
                output_ws = output_wb[self.module.MAIN_SHEET]
                self.assertEqual(source_ws["B13"].value, "-")
                self.assertIsNone(source_ws["Z13"].value)
                self.assertEqual(output_ws["B13"].value, "Y")
                self.assertEqual(output_ws["Z13"].value, "Verified current evidence; review before distribution.")
                self.assertEqual(output_ws["F13"].value, "Moderate")
                self.assertEqual(output_ws["AA13"].value, "=1+1")
                self.assertTrue(self.module.font_is_red(output_ws["B13"]))
                self.assertTrue(self.module.font_is_red(output_ws["Z13"]))
                self.assertTrue(self.module.validations_for_cell(output_ws, "B13"))
            finally:
                source_wb.close()
                output_wb.close()

    def test_write_copy_refuses_source_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "big-rocks.xlsx"
            manifest = tmp_path / "manifest.json"
            updates = tmp_path / "updates.json"
            current_source = tmp_path / "source-current.xlsx"
            stale_source = tmp_path / "source-stale.xlsx"
            self.write_shifted_main_workbook(workbook)
            self.write_source_workbook(current_source, updated="2026-06-20", notes="Automation Cloud.")
            self.write_source_workbook(stale_source, updated="2025-01-01", notes="Old evidence.")
            self.copy_fixture("source-manifest-v1.json", manifest)
            self.copy_fixture("proposed-updates-v1.json", updates)
            source_bytes = workbook.read_bytes()

            result = self.run_cli(
                workbook,
                manifest,
                "--row",
                "13",
                "--as-of-date",
                "2026-07-10",
                "--proposed-updates",
                str(updates),
                "--write-copy",
                str(workbook),
            )

            self.assertEqual(result.returncode, 2)
            self.assertEqual(json.loads(result.stderr)["error"]["code"], "in_place_write_refused")
            self.assertEqual(workbook.read_bytes(), source_bytes)


if __name__ == "__main__":
    unittest.main()
