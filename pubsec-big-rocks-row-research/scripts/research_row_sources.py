#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import re
import shutil
import sys
from copy import copy
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


SOURCE_MANIFEST_CONTRACT = "pubsec-big-rocks-row-research/source-manifest@1"
PROPOSED_UPDATES_CONTRACT = "pubsec-big-rocks-row-research/proposed-updates@1"
OUTPUT_CONTRACT = "pubsec-big-rocks-row-research/output@1"
ERROR_CONTRACT = "pubsec-big-rocks-row-research/error@1"

MAIN_SHEET = "PUBSEC Big Rocks_FY27"
MAX_HEADER_SCAN_ROWS = 100
MAX_SAFE_ROWS = 20_000
MAX_SAFE_COLUMNS = 500

ACCOUNT_HEADER_ALIASES = {
    "account",
    "account name",
    "accountname",
    "customer",
    "customer name",
    "uipath account",
}

TARGET_HEADER_ALIASES = {
    "Bot/License Utilization": {
        "bot/license utilization",
        "bot license utilization",
        "bot utilization",
        "license utilization",
        "utilization",
    },
    "Cloud Y/N": {
        "cloud y/n",
        "cloud yn",
        "cloud?",
        "cloud status",
        "on cloud",
    },
    "Consuming AI Units Today: Y/N": {
        "consuming ai units today: y/n",
        "consuming ai units today",
        "ai units consumed",
        "ai unit consumption",
        "using ai units",
    },
    "Agent Units Purchased Y/N": {
        "agent units purchased y/n",
        "agent units purchased",
        "purchased agent units",
        "agent unit entitlement",
    },
    "Test Status": {
        "test status",
        "test suite status",
        "test cloud status",
        "test manager status",
    },
    "IXP Status": {
        "ixp status",
        "document understanding status",
        "du/ixp status",
        "du ixp status",
    },
    "Agentic Status": {
        "agentic status",
        "agents status",
        "agent status",
    },
    "Regional Leader Only: Bell Curve Adoption Flag": {
        "regional leader only: bell curve adoption flag",
        "bell curve adoption flag",
        "adoption flag",
        "adoption segment",
    },
    "FY27 Big Rocks": {
        "fy27 big rocks",
        "fy 27 big rocks",
        "big rocks",
        "fy27 priorities",
    },
    "Tracking Value Realized": {
        "tracking value realized",
        "value realized tracking",
        "value tracking",
    },
    "At Risk/Churn Forecasted: Y/N": {
        "at risk/churn forecasted: y/n",
        "at risk churn forecasted",
        "risk/churn",
        "churn forecasted",
        "at risk",
    },
    "Notes / Evidence Additions": {
        "notes / evidence additions",
        "notes evidence additions",
        "notes/evidence",
        "evidence notes",
        "notes",
    },
}

DROPDOWN_VALUES = {
    "Bot/License Utilization": ("Low", "Moderate", "High", "Maxed"),
    "Cloud Y/N": ("Y", "N"),
    "Consuming AI Units Today: Y/N": ("Y", "N"),
    "Agent Units Purchased Y/N": ("Y", "N"),
    "Test Status": ("Not Yet", "Exploring", "PoC", "PRD", "Blocked"),
    "IXP Status": ("Not Yet", "Exploring", "PoC", "PRD", "Blocked"),
    "Agentic Status": ("Not Yet", "Exploring", "PoC", "PRD", "Blocked"),
    "Regional Leader Only: Bell Curve Adoption Flag": (
        "Early Adopter",
        "Early Majority",
        "Late Majority",
        "Laggard",
    ),
    "Tracking Value Realized": ("Don't Know", "Partially", "Yes", "Won't Share"),
    "At Risk/Churn Forecasted: Y/N": ("Y", "N"),
}

DATE_FIELD_HINTS = (
    "updated",
    "modified",
    "last activity",
    "last modified",
    "activity date",
    "report date",
    "export date",
    "as of",
)

TRUSTED_FRESHNESS_BASES = {
    "sharepoint_modified_at",
    "sfdc_last_modified_at",
    "slack_message_timestamp",
    "teams_message_timestamp",
    "onenote_modified_at",
    "workbook_report_date",
    "source_export_date",
}

MANIFEST_KEYS = {
    "contract_version",
    "data_classification",
    "retention_until",
    "sources",
}
SOURCE_KEYS = {
    "id",
    "kind",
    "path",
    "required",
    "source_updated_at",
    "freshness_basis",
}
UPDATES_KEYS = {"contract_version", "target", "updates"}
UPDATE_TARGET_KEYS = {"account", "row"}
UPDATE_KEYS = {"header", "value", "confidence", "evidence", "reason"}


class ContractError(Exception):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: Any | None = None,
        candidates: list[dict[str, Any]] | None = None,
        status: int = 2,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details
        self.candidates = candidates
        self.status = status

    def payload(self) -> dict[str, Any]:
        error: dict[str, Any] = {"code": self.code, "message": self.message}
        if self.details is not None:
            error["details"] = to_jsonable(self.details)
        if self.candidates is not None:
            error["candidates"] = to_jsonable(self.candidates)
        return {"contract_version": ERROR_CONTRACT, "error": error}


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def norm(value: Any) -> str:
    text = str(value or "").lower().replace("&", " and ")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(
        r"\b(inc|llc|ltd|corp|corporation|limited|department|dept|office|of|the|and|hq)\b",
        " ",
        text,
    )
    return re.sub(r"\s+", " ", text).strip()


def raw_norm(value: Any) -> str:
    return normalize_header(value)


ACCOUNT_ALIAS_KEYS = {normalize_header(alias) for alias in ACCOUNT_HEADER_ALIASES}
TARGET_ALIAS_INDEX: dict[str, str] = {}
for _canonical, _aliases in TARGET_HEADER_ALIASES.items():
    for _alias in {_canonical, *_aliases}:
        _key = normalize_header(_alias)
        if _key in TARGET_ALIAS_INDEX and TARGET_ALIAS_INDEX[_key] != _canonical:
            raise RuntimeError(f"Conflicting target header alias: {_alias}")
        TARGET_ALIAS_INDEX[_key] = _canonical


def canonical_target_header(value: Any) -> str | None:
    return TARGET_ALIAS_INDEX.get(normalize_header(value))


def clean(value: Any, limit: int = 500) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) > limit:
        return text[: limit - 1].rstrip() + "..."
    return text


def parse_as_of(value: str | None) -> datetime:
    if not value:
        return datetime.now(timezone.utc)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    raise ContractError("invalid_as_of_date", f"Invalid --as-of-date {value!r}; use YYYY-MM-DD.")


def parse_contract_datetime(value: Any, field: str) -> datetime:
    if not isinstance(value, str) or not value.strip():
        raise ContractError("invalid_contract", f"{field} must be a non-empty ISO-8601 date or timestamp.")
    text = value.strip()
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ContractError(
            "invalid_contract",
            f"{field} must be an ISO-8601 date or timestamp; got {value!r}.",
        ) from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def subtract_months(value: datetime, months: int) -> datetime:
    if months < 1:
        raise ContractError("invalid_recency_window", "--months must be at least 1.")
    month = value.month - months
    year = value.year
    while month <= 0:
        month += 12
        year -= 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return value.replace(year=year, month=month, day=day)


def file_modified_at(path: Path) -> datetime | None:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
    except OSError:
        return None


def format_dt(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def parse_dateish(value: Any) -> datetime | None:
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    patterns = [
        ("%Y-%m-%d", r"\b\d{4}-\d{1,2}-\d{1,2}\b"),
        ("%m/%d/%Y", r"\b\d{1,2}/\d{1,2}/\d{4}\b"),
        ("%m/%d/%y", r"\b\d{1,2}/\d{1,2}/\d{2}\b"),
        ("%m-%d-%Y", r"\b\d{1,2}-\d{1,2}-\d{4}\b"),
        ("%m-%d-%y", r"\b\d{1,2}-\d{1,2}-\d{2}\b"),
    ]
    for fmt, pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        try:
            return datetime.strptime(match.group(0), fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def sheet_date(ws_title: str) -> datetime | None:
    return parse_dateish(ws_title)


def row_update_dates(data: dict[str, Any]) -> list[dict[str, str]]:
    dates = []
    for key, value in data.items():
        key_l = key.lower()
        if any(hint in key_l for hint in DATE_FIELD_HINTS):
            parsed = parse_dateish(value)
            if parsed:
                dates.append({"field": key, "date": parsed.date().isoformat()})
    return dates


def blank_like(value: Any) -> bool:
    text = str(value or "").strip().replace("\r\n", "\n")
    compact = re.sub(r"\s+", "", text)
    return text == "" or compact in {"•", "••", "•••", "-", "--", "---"}


def read_json_object(path: Path, label: str) -> dict[str, Any]:
    if not path.exists():
        raise ContractError("file_not_found", f"{label} not found: {path}")
    if not path.is_file():
        raise ContractError("invalid_path", f"{label} must be a file: {path}")
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ContractError("invalid_json", f"Could not parse {label} {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise ContractError("invalid_contract", f"{label} root must be a JSON object.")
    return value


def reject_unknown_keys(value: dict[str, Any], allowed: set[str], location: str) -> None:
    unknown = sorted(set(value) - allowed)
    if unknown:
        raise ContractError(
            "invalid_contract",
            f"{location} contains unsupported fields: {', '.join(unknown)}.",
        )


def load_source_manifest(path: Path, as_of: datetime) -> dict[str, Any]:
    manifest_path = path.expanduser().resolve()
    manifest = read_json_object(manifest_path, "Source manifest")
    reject_unknown_keys(manifest, MANIFEST_KEYS, "Source manifest")
    if manifest.get("contract_version") != SOURCE_MANIFEST_CONTRACT:
        raise ContractError(
            "unsupported_manifest_version",
            f"Source manifest contract_version must be {SOURCE_MANIFEST_CONTRACT!r}.",
            details={"received": manifest.get("contract_version")},
        )
    classification = manifest.get("data_classification")
    if not isinstance(classification, str) or not classification.strip():
        raise ContractError("invalid_contract", "Source manifest data_classification is required.")
    retention_until = parse_contract_datetime(manifest.get("retention_until"), "retention_until")
    if retention_until.date() < as_of.date():
        raise ContractError(
            "retention_expired",
            "Source manifest retention_until is before the run date. Delete expired local data or create a newly authorized manifest.",
            details={"retention_until": retention_until.date().isoformat(), "as_of_date": as_of.date().isoformat()},
        )
    raw_sources = manifest.get("sources")
    if not isinstance(raw_sources, list):
        raise ContractError("invalid_contract", "Source manifest sources must be a JSON array.")

    sources: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_paths: set[Path] = set()
    for index, raw_source in enumerate(raw_sources):
        location = f"sources[{index}]"
        if not isinstance(raw_source, dict):
            raise ContractError("invalid_contract", f"{location} must be a JSON object.")
        reject_unknown_keys(raw_source, SOURCE_KEYS, location)
        source_id = raw_source.get("id")
        if not isinstance(source_id, str) or not re.fullmatch(r"[a-z0-9][a-z0-9._-]*", source_id):
            raise ContractError(
                "invalid_contract",
                f"{location}.id must use lowercase letters, digits, dots, underscores, or hyphens.",
            )
        if source_id in seen_ids:
            raise ContractError("invalid_contract", f"Duplicate source id: {source_id}")
        seen_ids.add(source_id)
        if raw_source.get("kind") != "xlsx":
            raise ContractError("invalid_contract", f"{location}.kind must be 'xlsx'.")
        raw_path = raw_source.get("path")
        if not isinstance(raw_path, str) or not raw_path.strip():
            raise ContractError("invalid_contract", f"{location}.path is required.")
        source_path = Path(raw_path).expanduser()
        if not source_path.is_absolute():
            source_path = manifest_path.parent / source_path
        source_path = source_path.resolve()
        if source_path.suffix.lower() != ".xlsx":
            raise ContractError("invalid_contract", f"{location}.path must reference an .xlsx file.")
        if source_path in seen_paths:
            raise ContractError("invalid_contract", f"Duplicate source path: {source_path}")
        seen_paths.add(source_path)
        required = raw_source.get("required", True)
        if not isinstance(required, bool):
            raise ContractError("invalid_contract", f"{location}.required must be true or false.")

        source_updated_at = None
        freshness_basis = raw_source.get("freshness_basis")
        if raw_source.get("source_updated_at") is not None:
            source_updated_at = parse_contract_datetime(
                raw_source.get("source_updated_at"),
                f"{location}.source_updated_at",
            )
            if freshness_basis not in TRUSTED_FRESHNESS_BASES:
                raise ContractError(
                    "invalid_contract",
                    f"{location}.freshness_basis must be one of: {', '.join(sorted(TRUSTED_FRESHNESS_BASES))}.",
                )
        elif freshness_basis is not None:
            raise ContractError(
                "invalid_contract",
                f"{location}.freshness_basis requires source_updated_at.",
            )

        sources.append(
            {
                "id": source_id,
                "kind": "xlsx",
                "path": source_path,
                "required": required,
                "source_updated_at": source_updated_at,
                "freshness_basis": freshness_basis,
            }
        )

    return {
        "path": manifest_path,
        "contract_version": SOURCE_MANIFEST_CONTRACT,
        "data_classification": classification.strip(),
        "retention_until": retention_until.date().isoformat(),
        "sources": sources,
    }


def detect_main_schema(ws) -> dict[str, Any]:
    candidates: list[dict[str, Any]] = []
    scan_rows = min(ws.max_row or 0, MAX_HEADER_SCAN_ROWS)
    scan_columns = min(ws.max_column or 0, MAX_SAFE_COLUMNS)
    for row in range(1, scan_rows + 1):
        account_columns: list[int] = []
        target_columns: dict[str, int] = {}
        duplicate_targets: dict[str, list[int]] = {}
        for col in range(1, scan_columns + 1):
            header = ws.cell(row, col).value
            key = normalize_header(header)
            if key in ACCOUNT_ALIAS_KEYS:
                account_columns.append(col)
            canonical = canonical_target_header(header)
            if canonical:
                if canonical in target_columns:
                    duplicate_targets.setdefault(canonical, [target_columns[canonical]]).append(col)
                else:
                    target_columns[canonical] = col
        if account_columns:
            candidates.append(
                {
                    "header_row": row,
                    "account_columns": account_columns,
                    "target_columns": target_columns,
                    "duplicate_targets": duplicate_targets,
                    "score": len(target_columns),
                }
            )
    if not candidates:
        raise ContractError(
            "main_schema_not_found",
            f"Could not find an account header within the first {scan_rows} rows of sheet {ws.title!r}.",
        )
    best_score = max(candidate["score"] for candidate in candidates)
    best = [candidate for candidate in candidates if candidate["score"] == best_score]
    if best_score == 0:
        raise ContractError(
            "main_schema_not_found",
            "An account header was found, but no recognized Big Rocks target headers were present.",
        )
    if len(best) != 1:
        raise ContractError(
            "ambiguous_main_header",
            "Multiple rows have equally strong Big Rocks header matches; identify a unique main header row before continuing.",
            candidates=[{"row": item["header_row"], "recognized_target_count": item["score"]} for item in best],
        )
    selected = best[0]
    if len(selected["account_columns"]) != 1 or selected["duplicate_targets"]:
        raise ContractError(
            "ambiguous_main_schema",
            "The detected main header contains duplicate account or target columns.",
            details={
                "header_row": selected["header_row"],
                "account_columns": selected["account_columns"],
                "duplicate_targets": selected["duplicate_targets"],
            },
        )
    selected["account_column"] = selected["account_columns"][0]
    selected["missing_target_headers"] = [
        header for header in TARGET_HEADER_ALIASES if header not in selected["target_columns"]
    ]
    return selected


def ensure_main_dimensions(ws) -> None:
    if (ws.max_row or 0) > MAX_SAFE_ROWS or (ws.max_column or 0) > MAX_SAFE_COLUMNS:
        raise ContractError(
            "workbook_dimensions_exceed_safety_limit",
            "Workbook dimensions exceed the safe deterministic scan limit.",
            details={
                "sheet": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "max_rows": MAX_SAFE_ROWS,
                "max_columns": MAX_SAFE_COLUMNS,
            },
        )


def account_match(query: Any, candidate: Any) -> tuple[str, float]:
    query_raw = raw_norm(query)
    candidate_raw = raw_norm(candidate)
    if query_raw and query_raw == candidate_raw:
        return "exact_normalized", 1.0
    query_norm = norm(query)
    candidate_norm = norm(candidate)
    if not query_norm or not candidate_norm:
        return "none", 0.0
    score = SequenceMatcher(None, query_norm, candidate_norm).ratio()
    contains = (
        len(query_norm) >= 8
        and len(candidate_norm) >= 8
        and (query_norm in candidate_norm or candidate_norm in query_norm)
    )
    if contains or score >= 0.78:
        return "approximate", score
    return "none", score


def load_main_row(workbook: Path, sheet: str, row: int | None, account: str | None) -> dict[str, Any]:
    try:
        wb = load_workbook(workbook, data_only=False, read_only=False)
    except Exception as exc:
        raise ContractError("workbook_open_failed", f"Could not open workbook {workbook}: {exc}") from exc
    try:
        if sheet not in wb.sheetnames:
            raise ContractError(
                "sheet_not_found",
                f"Sheet not found: {sheet}.",
                details={"available_sheets": wb.sheetnames},
            )
        ws = wb[sheet]
        ensure_main_dimensions(ws)
        schema = detect_main_schema(ws)
        account_col = schema["account_column"]
        target_row = row

        if target_row is not None:
            if target_row <= schema["header_row"] or target_row > ws.max_row:
                raise ContractError(
                    "invalid_target_row",
                    f"Target row {target_row} is outside the data rows for header row {schema['header_row']}.",
                )
            row_account = ws.cell(target_row, account_col).value
            if blank_like(row_account):
                raise ContractError("invalid_target_row", f"Target row {target_row} has no account value.")
            if account:
                match_type, _ = account_match(account, row_account)
                if match_type != "exact_normalized":
                    raise ContractError(
                        "account_row_mismatch",
                        "--row and --account do not resolve to the same exact normalized account.",
                        details={"row": target_row, "row_account": row_account, "requested_account": account},
                    )
        else:
            if not account:
                raise ContractError("target_required", "Provide --row or --account.")
            exact_matches: list[dict[str, Any]] = []
            approximate: list[dict[str, Any]] = []
            for candidate_row in range(schema["header_row"] + 1, ws.max_row + 1):
                name = ws.cell(candidate_row, account_col).value
                if blank_like(name):
                    continue
                match_type, score = account_match(account, name)
                candidate = {
                    "row": candidate_row,
                    "account": str(name),
                    "match_score": round(score, 3),
                }
                if match_type == "exact_normalized":
                    exact_matches.append(candidate)
                elif match_type == "approximate":
                    approximate.append(candidate)
            if len(exact_matches) > 1:
                raise ContractError(
                    "ambiguous_account",
                    "The account name resolves to multiple exact normalized rows. Use --row with the intended row and the exact account name.",
                    candidates=exact_matches,
                )
            if not exact_matches:
                approximate.sort(key=lambda item: (-item["match_score"], item["row"]))
                raise ContractError(
                    "account_not_found",
                    "No exact normalized account match was found. Fuzzy candidates are returned for review but are never selected automatically.",
                    candidates=approximate[:8],
                )
            target_row = exact_matches[0]["row"]

        assert target_row is not None
        headers = {col: ws.cell(schema["header_row"], col).value for col in range(1, ws.max_column + 1)}
        values: dict[str, Any] = {}
        for col, header in headers.items():
            if header is not None:
                values[str(header)] = ws.cell(target_row, col).value

        target_columns: dict[str, dict[str, Any]] = {}
        blanks: list[dict[str, Any]] = []
        for canonical, col in schema["target_columns"].items():
            cell = ws.cell(target_row, col)
            target_columns[canonical] = {
                "column": col,
                "coordinate": cell.coordinate,
                "workbook_header": headers.get(col),
            }
            if blank_like(cell.value):
                blanks.append({"cell": cell.coordinate, "header": canonical})

        account_value = ws.cell(target_row, account_col).value
        return {
            "sheet": sheet,
            "header_row": schema["header_row"],
            "row": target_row,
            "account": account_value,
            "account_column": account_col,
            "account_cell": ws.cell(target_row, account_col).coordinate,
            "target_columns": target_columns,
            "missing_target_headers": schema["missing_target_headers"],
            "headers": headers,
            "values": values,
            "blank_target_fields": blanks,
            "workbook_sheets": wb.sheetnames,
        }
    finally:
        wb.close()


def bounded_max_row(ws, cap: int = 1000) -> int:
    return min(ws.max_row or 0, cap)


def likely_header_row(ws, max_scan: int = 30) -> int | None:
    best_row = None
    best_score = -1
    for row in range(1, min(ws.max_row or 0, max_scan) + 1):
        values = [normalize_header(ws.cell(row, col).value) for col in range(1, min(ws.max_column, MAX_SAFE_COLUMNS) + 1)]
        account_score = sum(3 for value in values if value in ACCOUNT_ALIAS_KEYS)
        context_score = sum(1 for value in values if any(hint in value for hint in DATE_FIELD_HINTS) or value == "notes")
        score = account_score + context_score
        if account_score and score > best_score:
            best_row = row
            best_score = score
    return best_row


def row_to_dict(ws, header_row: int, row: int) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for col in range(1, min(ws.max_column, MAX_SAFE_COLUMNS) + 1):
        header = ws.cell(header_row, col).value
        if header is None:
            continue
        key = clean(header, 120)
        value = ws.cell(row, col).value
        if value is not None and str(value).strip() != "":
            data[key] = value
    return data


def find_name_col(ws, header_row: int) -> int | None:
    for col in range(1, min(ws.max_column, MAX_SAFE_COLUMNS) + 1):
        if normalize_header(ws.cell(header_row, col).value) in ACCOUNT_ALIAS_KEYS:
            return col
    for col in range(1, min(ws.max_column, MAX_SAFE_COLUMNS) + 1):
        value = normalize_header(ws.cell(header_row, col).value)
        if "account" in value and "owner" not in value and "id" not in value:
            return col
    return None


def record_freshness(
    data: dict[str, Any],
    ws_title: str,
    source: dict[str, Any] | None,
    cutoff: datetime,
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    update_dates = row_update_dates(data)
    parsed_row_dates = [
        (item, parse_dateish(item["date"]))
        for item in update_dates
        if parse_dateish(item["date"]) is not None
    ]
    effective: datetime | None = None
    basis: str | None = None
    if parsed_row_dates:
        selected, effective = max(parsed_row_dates, key=lambda item: item[1])
        basis = f"row_field:{selected['field']}"
    else:
        effective = sheet_date(ws_title)
        if effective:
            basis = "worksheet_title_date"
        elif source and source.get("source_updated_at"):
            effective = source["source_updated_at"]
            basis = f"manifest:{source['freshness_basis']}"
    status = "undated" if effective is None else ("current" if effective >= cutoff else "stale")
    return (
        {
            "status": status,
            "effective_date": effective.date().isoformat() if effective else None,
            "basis": basis,
            "cutoff_date": cutoff.date().isoformat(),
            "local_file_modified_at": None,
            "local_file_mtime_is_eligibility_evidence": False,
        },
        update_dates,
    )


def route_evidence(
    record: dict[str, Any],
    fill_eligible: list[dict[str, Any]],
    discovery_leads: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    include_stale: bool,
) -> None:
    freshness_status = record["freshness"]["status"]
    exact_account = record["match_type"] == "exact_normalized"
    if freshness_status == "current" and exact_account:
        record["fill_eligible"] = True
        fill_eligible.append(record)
        return

    record["fill_eligible"] = False
    if not exact_account:
        reason = "account match is not exact"
    elif freshness_status == "stale":
        reason = "evidence date is before the cutoff"
    else:
        reason = "evidence has no trusted freshness date"
    record["discovery_reason"] = reason
    if freshness_status == "current" or include_stale:
        discovery_leads.append(record)
    else:
        excluded.append(
            {
                "evidence_id": record["evidence_id"],
                "source_file": record["source_file"],
                "sheet": record["sheet"],
                "row": record["row"],
                "matched_name": record.get("matched_name"),
                "reason": reason,
                "freshness": record["freshness"],
            }
        )


def match_records(
    source: dict[str, Any],
    account: str,
    cutoff: datetime,
    fill_eligible: list[dict[str, Any]],
    discovery_leads: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    include_stale: bool = False,
    max_rows_per_sheet: int = 12,
) -> None:
    source_path: Path = source["path"]
    try:
        wb = load_workbook(source_path, data_only=True, read_only=False)
    except Exception as exc:
        raise ContractError("source_open_failed", f"Could not open source {source_path}: {exc}") from exc
    try:
        source_mtime = file_modified_at(source_path)
        for ws in wb.worksheets:
            header_row = likely_header_row(ws)
            if header_row is None:
                continue
            name_col = find_name_col(ws, header_row)
            if not name_col:
                continue
            sheet_records: list[dict[str, Any]] = []
            for row in range(header_row + 1, bounded_max_row(ws) + 1):
                name = ws.cell(row, name_col).value
                if blank_like(name):
                    continue
                match_type, score = account_match(account, name)
                if match_type == "none":
                    continue
                data = row_to_dict(ws, header_row, row)
                freshness, update_dates = record_freshness(data, ws.title, source, cutoff)
                freshness["local_file_modified_at"] = format_dt(source_mtime)
                sheet_records.append(
                    {
                        "evidence_id": f"{source['id']}:{ws.title}:{row}",
                        "scope": "manifest_source",
                        "source_id": source["id"],
                        "source_file": str(source_path),
                        "sheet": ws.title,
                        "row": row,
                        "matched_name": str(name),
                        "match_type": match_type,
                        "match_score": round(score, 3),
                        "freshness": freshness,
                        "row_update_dates": update_dates,
                        "data": data,
                    }
                )
            sheet_records.sort(key=lambda item: (-item["match_score"], item["row"]))
            for record in sheet_records[:max_rows_per_sheet]:
                route_evidence(record, fill_eligible, discovery_leads, excluded, include_stale)
    finally:
        wb.close()


def scan_workbook_tabs(
    workbook: Path,
    account: str,
    main_sheet: str,
    cutoff: datetime,
    fill_eligible: list[dict[str, Any]],
    discovery_leads: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    include_stale: bool = False,
) -> None:
    try:
        wb = load_workbook(workbook, data_only=True, read_only=False)
    except Exception as exc:
        raise ContractError("workbook_open_failed", f"Could not scan workbook tabs in {workbook}: {exc}") from exc
    try:
        source_mtime = file_modified_at(workbook)
        for ws in wb.worksheets:
            if ws.title == main_sheet:
                continue
            header_row = likely_header_row(ws, max_scan=40)
            name_col = find_name_col(ws, header_row) if header_row else None
            if header_row and name_col:
                for row in range(header_row + 1, bounded_max_row(ws) + 1):
                    name = ws.cell(row, name_col).value
                    if blank_like(name):
                        continue
                    match_type, score = account_match(account, name)
                    if match_type == "none":
                        continue
                    data = row_to_dict(ws, header_row, row)
                    freshness, update_dates = record_freshness(data, ws.title, None, cutoff)
                    freshness["local_file_modified_at"] = format_dt(source_mtime)
                    record = {
                        "evidence_id": f"main-workbook:{ws.title}:{row}",
                        "scope": "internal_workbook",
                        "source_id": "main-workbook",
                        "source_file": str(workbook),
                        "sheet": ws.title,
                        "row": row,
                        "matched_name": str(name),
                        "match_type": match_type,
                        "match_score": round(score, 3),
                        "freshness": freshness,
                        "row_update_dates": update_dates,
                        "data": data,
                    }
                    route_evidence(record, fill_eligible, discovery_leads, excluded, include_stale)
                continue

            for row in range(1, bounded_max_row(ws) + 1):
                row_values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, MAX_SAFE_COLUMNS) + 1)]
                row_text = " ".join(str(value) for value in row_values if value is not None)
                if not row_text or norm(account) not in norm(row_text):
                    continue
                data = {"row_text": clean(row_text, 900)}
                freshness, update_dates = record_freshness(data, ws.title, None, cutoff)
                freshness["local_file_modified_at"] = format_dt(source_mtime)
                record = {
                    "evidence_id": f"main-workbook:{ws.title}:{row}",
                    "scope": "internal_workbook",
                    "source_id": "main-workbook",
                    "source_file": str(workbook),
                    "sheet": ws.title,
                    "row": row,
                    "matched_name": None,
                    "match_type": "unstructured_mention",
                    "match_score": None,
                    "freshness": freshness,
                    "row_update_dates": update_dates,
                    "data": data,
                }
                route_evidence(record, fill_eligible, discovery_leads, excluded, include_stale)
    finally:
        wb.close()


def summarize_recommendation_leads(records: list[dict[str, Any]]) -> dict[str, list[str]]:
    text_chunks = []
    for record in records:
        for key, value in record.get("data", {}).items():
            if value is not None:
                text_chunks.append(f"{key}: {value}")
    text = "\n".join(text_chunks).lower()

    leads: dict[str, list[str]] = {}
    if any(token in text for token in ["automation cloud", "acps", "govcloud", "active public sector cloud"]):
        leads.setdefault("Cloud Y/N", []).append(
            "Current account-specific cloud evidence appears; confirm current versus target platform before proposing a value."
        )
    if any(token in text for token in ["on premise", "on-premise", "msi", "automation suite", "studio only"]):
        leads.setdefault("Cloud Y/N", []).append(
            "Current account-specific on-premises evidence appears; do not mark cloud Y for a target-state migration."
        )
    if any(token in text for token in ["using du", "implementing du", "document understanding", "communications mining", "ixp"]):
        leads.setdefault("IXP Status", []).append(
            "Current DU/IXP evidence appears; classify the stage only from explicit source language."
        )
    if "using du" in text:
        leads.setdefault("Consuming AI Units Today: Y/N", []).append(
            "Current active DU language appears; verify AI Unit telemetry before proposing Y."
        )
    if any(token in text for token in ["agentic", "autopilot", "agents"]):
        leads.setdefault("Agentic Status", []).append(
            "Current agentic language appears; classify the stage only from explicit source language."
        )
    if any(token in text for token in ["risk/downsell", "customer status: churn", "ischurn: churn", "program has been put on hold", "high risk"]):
        leads.setdefault("At Risk/Churn Forecasted: Y/N", []).append(
            "Current risk language appears; verify it against the account-specific forecast source."
        )
    return leads


def load_proposed_updates(path: Path) -> dict[str, Any]:
    updates_path = path.expanduser().resolve()
    proposal = read_json_object(updates_path, "Proposed updates")
    reject_unknown_keys(proposal, UPDATES_KEYS, "Proposed updates")
    if proposal.get("contract_version") != PROPOSED_UPDATES_CONTRACT:
        raise ContractError(
            "unsupported_updates_version",
            f"Proposed updates contract_version must be {PROPOSED_UPDATES_CONTRACT!r}.",
            details={"received": proposal.get("contract_version")},
        )
    target = proposal.get("target")
    if not isinstance(target, dict):
        raise ContractError("invalid_contract", "Proposed updates target must be a JSON object.")
    reject_unknown_keys(target, UPDATE_TARGET_KEYS, "Proposed updates target")
    if not isinstance(target.get("account"), str) or not target["account"].strip():
        raise ContractError("invalid_contract", "Proposed updates target.account is required.")
    if not isinstance(target.get("row"), int) or target["row"] < 1:
        raise ContractError("invalid_contract", "Proposed updates target.row must be a positive integer.")
    updates = proposal.get("updates")
    if not isinstance(updates, list) or not updates:
        raise ContractError("invalid_contract", "Proposed updates updates must be a non-empty JSON array.")
    for index, update in enumerate(updates):
        location = f"updates[{index}]"
        if not isinstance(update, dict):
            raise ContractError("invalid_contract", f"{location} must be a JSON object.")
        reject_unknown_keys(update, UPDATE_KEYS, location)
        if not isinstance(update.get("header"), str) or not update["header"].strip():
            raise ContractError("invalid_contract", f"{location}.header is required.")
        if not isinstance(update.get("value"), str) or blank_like(update["value"]):
            raise ContractError("invalid_contract", f"{location}.value must be a non-blank string.")
        confidence = update.get("confidence")
        if confidence not in {"High", "Medium"}:
            raise ContractError("invalid_contract", f"{location}.confidence must be High or Medium.")
        evidence = update.get("evidence")
        if not isinstance(evidence, list) or not evidence or not all(isinstance(item, str) and item.strip() for item in evidence):
            raise ContractError("invalid_contract", f"{location}.evidence must be a non-empty array of evidence IDs.")
        if update.get("reason") is not None and not isinstance(update["reason"], str):
            raise ContractError("invalid_contract", f"{location}.reason must be a string when provided.")
    proposal["path"] = updates_path
    return proposal


def validations_for_cell(ws, coordinate: str) -> list[Any]:
    validations = []
    collection = getattr(ws, "data_validations", None)
    if not collection:
        return validations
    for validation in collection.dataValidation:
        try:
            if coordinate in validation.ranges:
                validations.append(validation)
        except (TypeError, ValueError):
            continue
    return validations


def resolve_validation_values(wb, validation) -> tuple[list[str] | None, str | None]:
    formula = str(validation.formula1 or "").strip()
    if not formula:
        return None, "list validation has no formula1"
    if formula.startswith('"') and formula.endswith('"'):
        return [item.strip() for item in formula[1:-1].replace('""', '"').split(",")], None
    expression = formula[1:] if formula.startswith("=") else formula
    match = re.fullmatch(r"(?:'((?:[^']|'')+)'|([^!]+))!(\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)", expression)
    if not match:
        return None, f"unsupported list validation formula: {formula}"
    sheet_name = (match.group(1) or match.group(2)).replace("''", "'")
    if sheet_name not in wb.sheetnames:
        return None, f"validation references missing sheet: {sheet_name}"
    min_col, min_row, max_col, max_row = range_boundaries(match.group(3))
    values = []
    ref_ws = wb[sheet_name]
    for row in ref_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return values, None


def validate_dropdown(wb, ws, coordinate: str, canonical_header: str, value: str) -> dict[str, Any]:
    errors: list[str] = []
    contract_allowed = list(DROPDOWN_VALUES.get(canonical_header, ()))
    if contract_allowed and value not in contract_allowed:
        errors.append(
            f"value {value!r} is not in the contract dropdown: {', '.join(contract_allowed)}"
        )
    workbook_allowed_sets: list[list[str]] = []
    for validation in validations_for_cell(ws, coordinate):
        if validation.type != "list":
            continue
        allowed, error = resolve_validation_values(wb, validation)
        if error:
            errors.append(error)
            continue
        assert allowed is not None
        workbook_allowed_sets.append(allowed)
        if value not in allowed:
            errors.append(
                f"value {value!r} is not allowed by the workbook dropdown: {', '.join(allowed)}"
            )
    return {
        "valid": not errors,
        "contract_allowed_values": contract_allowed,
        "workbook_allowed_values": workbook_allowed_sets,
        "errors": errors,
    }


def preview_proposed_updates(
    workbook: Path,
    target_row: dict[str, Any],
    proposal: dict[str, Any],
    fill_eligible_evidence_ids: set[str],
) -> dict[str, Any]:
    proposal_target = proposal["target"]
    target_errors = []
    if proposal_target["row"] != target_row["row"]:
        target_errors.append(
            f"proposal row {proposal_target['row']} does not match resolved row {target_row['row']}"
        )
    match_type, _ = account_match(proposal_target["account"], target_row["account"])
    if match_type != "exact_normalized":
        target_errors.append(
            f"proposal account {proposal_target['account']!r} does not exactly match {target_row['account']!r}"
        )

    try:
        wb = load_workbook(workbook, data_only=False, read_only=False)
    except Exception as exc:
        raise ContractError("workbook_open_failed", f"Could not preview updates for {workbook}: {exc}") from exc
    try:
        ws = wb[target_row["sheet"]]
        entries = []
        seen_headers: set[str] = set()
        for update in proposal["updates"]:
            errors: list[str] = list(target_errors)
            ineligible_evidence = [
                evidence_id
                for evidence_id in update["evidence"]
                if evidence_id not in fill_eligible_evidence_ids
            ]
            evidence_ids_fill_eligible = not ineligible_evidence
            if ineligible_evidence:
                errors.append(
                    "evidence IDs are not fill-eligible in this run: "
                    + ", ".join(ineligible_evidence)
                )
            canonical = canonical_target_header(update["header"])
            if not canonical:
                errors.append(f"unrecognized target header: {update['header']}")
            elif canonical in seen_headers:
                errors.append(f"duplicate proposed update for {canonical}")
            else:
                seen_headers.add(canonical)

            column_info = target_row["target_columns"].get(canonical) if canonical else None
            coordinate = column_info["coordinate"] if column_info else None
            current_value = None
            formula_safe = False
            existing_value_blank = False
            dropdown = {"valid": False, "contract_allowed_values": [], "workbook_allowed_values": [], "errors": []}
            if canonical and not column_info:
                errors.append(f"workbook has no recognized column for {canonical}")
            elif coordinate:
                cell = ws[coordinate]
                current_value = cell.value
                formula_safe = cell.data_type != "f" and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                )
                if not formula_safe:
                    errors.append("target cell contains a formula")
                existing_value_blank = blank_like(cell.value)
                if not existing_value_blank:
                    errors.append("target cell contains a substantive existing value")
                dropdown = validate_dropdown(wb, ws, coordinate, canonical, update["value"])
                errors.extend(dropdown["errors"])

            entries.append(
                {
                    "header": canonical or update["header"],
                    "requested_header": update["header"],
                    "workbook_header": column_info.get("workbook_header") if column_info else None,
                    "cell": coordinate,
                    "current_value": current_value,
                    "proposed_value": update["value"],
                    "confidence": update["confidence"],
                    "evidence": update["evidence"],
                    "reason": update.get("reason"),
                    "checks": {
                        "column_resolved": bool(column_info),
                        "formula_safe": formula_safe,
                        "existing_value_blank_or_placeholder": existing_value_blank,
                        "dropdown_valid": dropdown["valid"],
                        "evidence_ids_fill_eligible": evidence_ids_fill_eligible,
                    },
                    "dropdown_validation": dropdown,
                    "valid": not errors,
                    "errors": errors,
                }
            )
        return {
            "contract_version": PROPOSED_UPDATES_CONTRACT,
            "proposal_path": str(proposal["path"]),
            "target": {
                "sheet": target_row["sheet"],
                "row": target_row["row"],
                "account": target_row["account"],
            },
            "valid": all(entry["valid"] for entry in entries),
            "proposed_changes": entries,
            "source_workbook_modified": False,
        }
    finally:
        wb.close()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ensure_snapshot_dimensions(wb) -> None:
    for ws in wb.worksheets:
        if (ws.max_row or 0) > MAX_SAFE_ROWS or (ws.max_column or 0) > MAX_SAFE_COLUMNS:
            raise ContractError(
                "copy_verification_limit_exceeded",
                "A worksheet exceeds the full-copy verification limits; no output copy was written.",
                details={"sheet": ws.title, "rows": ws.max_row, "columns": ws.max_column},
            )


def workbook_value_snapshot(wb) -> dict[tuple[str, str], dict[str, Any]]:
    ensure_snapshot_dimensions(wb)
    snapshot: dict[tuple[str, str], dict[str, Any]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=max(ws.max_row or 1, 1),
            min_col=1,
            max_col=max(ws.max_column or 1, 1),
        ):
            for cell in row:
                if cell.value is not None:
                    snapshot[(ws.title, cell.coordinate)] = {
                        "value": to_jsonable(cell.value),
                        "data_type": cell.data_type,
                    }
    return snapshot


def formula_snapshot(wb) -> dict[tuple[str, str], Any]:
    return {
        key: value["value"]
        for key, value in workbook_value_snapshot(wb).items()
        if value["data_type"] == "f"
    }


def validation_snapshot(wb) -> list[dict[str, Any]]:
    values = []
    attributes = (
        "type",
        "formula1",
        "formula2",
        "sqref",
        "operator",
        "allow_blank",
        "showDropDown",
        "showErrorMessage",
        "error",
        "errorTitle",
        "showInputMessage",
        "prompt",
        "promptTitle",
    )
    for ws in wb.worksheets:
        collection = getattr(ws, "data_validations", None)
        if not collection:
            continue
        for validation in collection.dataValidation:
            values.append(
                {
                    "sheet": ws.title,
                    **{attribute: str(getattr(validation, attribute, None)) for attribute in attributes},
                }
            )
    return sorted(values, key=lambda item: json.dumps(item, sort_keys=True))


def font_is_red(cell) -> bool:
    color = cell.font.color
    return bool(color and color.type == "rgb" and str(color.rgb).upper() in {"FFFF0000", "00FF0000"})


def write_verified_copy(
    source_workbook: Path,
    destination_value: str,
    preview: dict[str, Any],
) -> dict[str, Any]:
    if not preview["valid"]:
        raise ContractError("invalid_update_preview", "Proposed updates failed preview validation; no copy was written.", status=3)
    source = source_workbook.resolve()
    destination = Path(destination_value).expanduser().resolve()
    if source == destination:
        raise ContractError("in_place_write_refused", "--write-copy must not resolve to the source workbook path.")
    if destination.suffix.lower() != ".xlsx":
        raise ContractError("invalid_output_path", "--write-copy must use an .xlsx path.")
    if destination.exists():
        raise ContractError("output_exists", f"Refusing to overwrite existing output copy: {destination}")
    if not destination.parent.exists():
        raise ContractError("output_directory_missing", f"Output directory does not exist: {destination.parent}")

    source_hash_before = sha256_file(source)
    try:
        source_wb = load_workbook(source, data_only=False, read_only=False)
        try:
            before_values = workbook_value_snapshot(source_wb)
            before_formulas = formula_snapshot(source_wb)
            before_validations = validation_snapshot(source_wb)
        finally:
            source_wb.close()

        shutil.copy2(source, destination)
        output_wb = load_workbook(destination, data_only=False, read_only=False)
        try:
            ws = output_wb[preview["target"]["sheet"]]
            for change in preview["proposed_changes"]:
                cell = ws[change["cell"]]
                cell.value = change["proposed_value"]
                red_font = copy(cell.font)
                red_font.color = "FFFF0000"
                cell.font = red_font
            output_wb.save(destination)
        finally:
            output_wb.close()

        verified_wb = load_workbook(destination, data_only=False, read_only=False)
        try:
            after_values = workbook_value_snapshot(verified_wb)
            after_formulas = formula_snapshot(verified_wb)
            after_validations = validation_snapshot(verified_wb)
            expected_keys = {
                (preview["target"]["sheet"], change["cell"]): change
                for change in preview["proposed_changes"]
            }
            existing_values_preserved = all(
                after_values.get(key) == value
                for key, value in before_values.items()
                if key not in expected_keys
            )
            unexpected_new_cells = [
                key for key in after_values if key not in before_values and key not in expected_keys
            ]
            changed_values_valid = all(
                after_values.get(key, {}).get("value") == to_jsonable(change["proposed_value"])
                for key, change in expected_keys.items()
            )
            target_ws = verified_wb[preview["target"]["sheet"]]
            red_font_verified = all(font_is_red(target_ws[change["cell"]]) for change in preview["proposed_changes"])
            dropdown_values_valid = all(
                validate_dropdown(
                    verified_wb,
                    target_ws,
                    change["cell"],
                    change["header"],
                    change["proposed_value"],
                )["valid"]
                for change in preview["proposed_changes"]
            )
        finally:
            verified_wb.close()

        source_unchanged = sha256_file(source) == source_hash_before
        checks = {
            "source_unchanged": source_unchanged,
            "existing_values_preserved": existing_values_preserved and not unexpected_new_cells,
            "formulas_preserved": before_formulas == after_formulas,
            "data_validations_preserved": before_validations == after_validations,
            "changed_values_valid": changed_values_valid,
            "dropdown_values_valid": dropdown_values_valid,
            "red_font_verified": red_font_verified,
        }
        if not all(checks.values()):
            raise ContractError(
                "copy_verification_failed",
                "The saved workbook copy failed post-write verification and was removed.",
                details={**checks, "unexpected_new_cells": unexpected_new_cells},
            )
        return {
            "output_path": str(destination),
            "source_path": str(source),
            "verified": True,
            "checks": checks,
            "changed_cells": [change["cell"] for change in preview["proposed_changes"]],
        }
    except Exception:
        if destination.exists():
            destination.unlink()
        raise


def to_jsonable(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): to_jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def render_markdown(result: dict[str, Any]) -> str:
    row = result["target_row"]
    evidence = result["evidence"]
    lines = [
        "**Account**",
        f"Row: {row['row']}",
        f"Account: {row['account']}",
        f"Header row: {row['header_row']}",
        f"Recency cutoff: {result['recency']['cutoff_date']}",
        f"As of: {result['recency']['as_of_date']}",
        f"Output contract: {result['contract_version']}",
        "",
        "**Do Not Fill Guidance**",
        "- Only records under Fill-Eligible Current Evidence can inform proposed cell values.",
        "- Discovery leads, stale evidence, undated evidence, fuzzy account matches, missing sources, formulas, and existing substantive values are fail-closed conditions.",
        "",
        "**Blank / Placeholder Target Fields**",
    ]
    for blank in row["blank_target_fields"]:
        lines.append(f"- {blank['cell']} {blank['header']}")
    if not row["blank_target_fields"]:
        lines.append("- None")

    lines += ["", "**Fill-Eligible Current Evidence**"]
    for record in evidence["fill_eligible_current"]:
        lines.append(
            f"- {Path(record['source_file']).name} / {record['sheet']} row {record['row']} "
            f"matched `{record['matched_name']}` ({record['freshness']['basis']}={record['freshness']['effective_date']})"
        )
    if not evidence["fill_eligible_current"]:
        lines.append("- None")

    lines += ["", "**Discovery Leads (Never Fill Directly)**"]
    for record in evidence["discovery_leads"]:
        lines.append(
            f"- {Path(record['source_file']).name} / {record['sheet']} row {record['row']}: "
            f"{record['discovery_reason']}"
        )
    if not evidence["discovery_leads"]:
        lines.append("- None")

    lines += ["", "**Recommendation Leads To Investigate**"]
    for header, leads in result["recommendation_leads"].items():
        for lead in leads:
            lines.append(f"- {header}: {lead}")
    if not result["recommendation_leads"]:
        lines.append("- None from fill-eligible current evidence; continue account-specific source research.")

    lines += ["", "**Missing Optional Source Files**"]
    if result["missing_optional_source_files"]:
        for item in result["missing_optional_source_files"]:
            lines.append(f"- {item['id']}: {item['path']}")
    else:
        lines.append("- None")

    lines += ["", "**Excluded Evidence**"]
    if evidence["excluded"]:
        for item in evidence["excluded"][:40]:
            lines.append(f"- {item['evidence_id']}: {item['reason']}")
    else:
        lines.append("- None")

    preview = result.get("proposed_update_preview")
    if preview:
        lines += ["", "**Proposed Update Preview**", f"- Valid: {preview['valid']}"]
        for change in preview["proposed_changes"]:
            status = "valid" if change["valid"] else "blocked: " + "; ".join(change["errors"])
            lines.append(f"- {change['cell'] or change['requested_header']}: {change['proposed_value']} ({status})")
    write_copy = result.get("write_copy")
    if write_copy:
        lines += ["", "**Verified Local Copy**", f"- {write_copy['output_path']}"]
    return "\n".join(lines)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Research one PubSec Big Rocks workbook row against explicitly manifested local sources."
    )
    parser.add_argument("--workbook", required=True, help="Path to the source Big Rocks .xlsx workbook")
    parser.add_argument("--manifest", help=f"Path to a {SOURCE_MANIFEST_CONTRACT} JSON source manifest")
    parser.add_argument("--sheet", default=MAIN_SHEET, help="Workbook sheet name")
    parser.add_argument("--row", type=int, help="Target worksheet row")
    parser.add_argument("--account", help="Exact normalized target account name")
    parser.add_argument("--months", type=int, default=3, help="Calendar-month recency window; default is 3")
    parser.add_argument("--max-age-days", type=int, help="Optional positive day-based recency window override")
    parser.add_argument("--as-of-date", help="Run date for cutoff calculation, YYYY-MM-DD. Defaults to today.")
    parser.add_argument(
        "--include-stale",
        action="store_true",
        help="Return stale or undated matches only under discovery_leads; never as fill-eligible evidence",
    )
    parser.add_argument("--proposed-updates", help=f"Path to a {PROPOSED_UPDATES_CONTRACT} JSON file")
    parser.add_argument("--write-copy", help="Write validated proposed updates to a new local .xlsx copy")
    parser.add_argument("--format", choices=["json", "markdown"], default="json")
    parser.add_argument("--source", action="append", help=argparse.SUPPRESS)
    parser.add_argument("--sources-only", action="store_true", help=argparse.SUPPRESS)
    return parser


def run(args: argparse.Namespace) -> tuple[dict[str, Any], int]:
    if args.source or args.sources_only:
        raise ContractError(
            "legacy_source_flags_removed",
            f"Legacy --source and --sources-only inputs are disabled. Put every explicit source in a {SOURCE_MANIFEST_CONTRACT} JSON file and pass --manifest.",
        )
    if not args.manifest:
        raise ContractError(
            "manifest_required",
            f"--manifest is required. Create a {SOURCE_MANIFEST_CONTRACT} JSON file; implicit machine-local source discovery is no longer supported.",
        )
    if args.max_age_days is not None and args.max_age_days < 1:
        raise ContractError("invalid_recency_window", "--max-age-days must be at least 1.")
    if args.write_copy and not args.proposed_updates:
        raise ContractError("proposed_updates_required", "--write-copy requires --proposed-updates.")

    as_of = parse_as_of(args.as_of_date)
    cutoff = (
        as_of - timedelta(days=args.max_age_days)
        if args.max_age_days is not None
        else subtract_months(as_of, args.months)
    )
    workbook = Path(args.workbook).expanduser().resolve()
    if not workbook.exists():
        raise ContractError("workbook_not_found", f"Workbook not found: {workbook}")
    if workbook.suffix.lower() != ".xlsx":
        raise ContractError("unsupported_workbook_type", "The source workbook must be an .xlsx file.")

    manifest = load_source_manifest(Path(args.manifest), as_of)
    missing_required = [source for source in manifest["sources"] if source["required"] and not source["path"].exists()]
    if missing_required:
        raise ContractError(
            "required_sources_missing",
            "One or more required manifest sources are missing. Restore them or explicitly mark them optional after reviewing the coverage impact.",
            details=[{"id": source["id"], "path": str(source["path"])} for source in missing_required],
        )
    missing_optional = [
        {"id": source["id"], "path": str(source["path"])}
        for source in manifest["sources"]
        if not source["required"] and not source["path"].exists()
    ]

    target_row = load_main_row(workbook, args.sheet, args.row, args.account)
    account = str(target_row["account"])
    fill_eligible: list[dict[str, Any]] = []
    discovery_leads: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for source in manifest["sources"]:
        if source["path"].exists():
            match_records(
                source,
                account,
                cutoff,
                fill_eligible,
                discovery_leads,
                excluded,
                include_stale=args.include_stale,
            )
    scan_workbook_tabs(
        workbook,
        account,
        args.sheet,
        cutoff,
        fill_eligible,
        discovery_leads,
        excluded,
        include_stale=args.include_stale,
    )

    preview = None
    write_copy = None
    status = 0
    if args.proposed_updates:
        proposal = load_proposed_updates(Path(args.proposed_updates))
        preview = preview_proposed_updates(
            workbook,
            target_row,
            proposal,
            {record["evidence_id"] for record in fill_eligible},
        )
        if preview["valid"] and args.write_copy:
            write_copy = write_verified_copy(workbook, args.write_copy, preview)
        elif not preview["valid"]:
            status = 3

    result = {
        "contract_version": OUTPUT_CONTRACT,
        "recency": {
            "as_of_date": as_of.date().isoformat(),
            "cutoff_date": cutoff.date().isoformat(),
            "months": args.months,
            "max_age_days": args.max_age_days,
            "eligibility_rule": "Use the most specific trusted row, worksheet-title, or manifest upstream timestamp. Local file modification time is diagnostic only.",
        },
        "source_manifest": {
            "contract_version": manifest["contract_version"],
            "path": str(manifest["path"]),
            "data_classification": manifest["data_classification"],
            "retention_until": manifest["retention_until"],
            "source_count": len(manifest["sources"]),
        },
        "target_row": target_row,
        "evidence": {
            "fill_eligible_current": fill_eligible,
            "discovery_leads": discovery_leads,
            "excluded": excluded,
        },
        "recommendation_leads": summarize_recommendation_leads(fill_eligible),
        "missing_optional_source_files": missing_optional,
        "proposed_update_preview": preview,
        "write_copy": write_copy,
    }
    return to_jsonable(result), status


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        result, status = run(args)
    except ContractError as exc:
        print(json.dumps(exc.payload(), indent=2), file=sys.stderr)
        return exc.status
    if args.format == "markdown":
        print(render_markdown(result))
    else:
        print(json.dumps(result, indent=2))
    return status


if __name__ == "__main__":
    raise SystemExit(main())
